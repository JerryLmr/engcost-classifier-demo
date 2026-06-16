import io
from collections import Counter
from pathlib import Path
from typing import Any, Dict, List

import openpyxl
from fastapi import HTTPException, UploadFile


REQUIRED_RESULT_COLUMNS = [
    "工程名称",
    "catalog_id",
    "一级分类",
    "二级分类",
    "维修状态",
    "标准对象",
    "是否复合工程",
    "复合候选目录",
    "是否紧急维修",
    "是否白蚁相关",
    "是否建议复核",
    "候选目录",
    "分类依据",
]

FOCUS_SAMPLE_LIMIT = 2000
OUT_OF_SCOPE_ID = "OUT_OF_SCOPE"


def _split_pipe_text(value: Any) -> List[str]:
    text = str(value or "").strip()
    if not text:
        return []
    return [part.strip() for part in text.split("|") if part.strip()]


def _is_yes(value: Any) -> bool:
    return value is True or str(value or "").strip() == "是"


def _method_for(catalog_id: Any) -> str:
    return "体系外默认分类" if str(catalog_id or "").strip() == OUT_OF_SCOPE_ID else "LLM主分类"


def _structure_type(is_composite: bool) -> str:
    return "composite_project" if is_composite else "single_project"


def _read_result_rows_from_workbook(workbook: openpyxl.Workbook, source_name: str) -> List[Dict[str, Any]]:
    worksheet = workbook.active
    rows = worksheet.iter_rows(values_only=True)
    header = list(next(rows))
    index = {name: i for i, name in enumerate(header)}

    missing = [column for column in REQUIRED_RESULT_COLUMNS if column not in index]
    if missing:
        raise HTTPException(status_code=400, detail=f"结果文件缺少必要列: {', '.join(missing)}")

    records: List[Dict[str, Any]] = []
    for row_num, row in enumerate(rows, start=2):
        if not any(row):
            continue
        project_name = row[index["工程名称"]]
        if project_name is None or str(project_name).strip() == "":
            continue

        catalog_id = row[index["catalog_id"]]
        is_composite = _is_yes(row[index["是否复合工程"]])
        secondary_candidates = _split_pipe_text(row[index["复合候选目录"]])
        method = _method_for(catalog_id)
        records.append(
            {
                "source_file": source_name,
                "row_num": row_num,
                "project_name": str(project_name),
                "catalog_id": catalog_id,
                "level1": row[index["一级分类"]],
                "level2": row[index["二级分类"]],
                "level3_item": row[index["二级分类"]],
                "matched_level3_items": [],
                "method": method,
                "confidence": "",
                "match_type": "out_of_scope" if method == "体系外默认分类" else "standard_catalog",
                "reason": row[index["分类依据"]],
                "needs_review": _is_yes(row[index["是否建议复核"]]),
                "candidate_ids": [],
                "candidate_labels": _split_pipe_text(row[index["候选目录"]]),
                "candidate_level3_items": [],
                "is_composite": is_composite,
                "structure_type": _structure_type(is_composite),
                "composite_reason": "疑似复合工程" if is_composite else "",
                "secondary_candidates": secondary_candidates,
            }
        )
    return records


def load_records_from_upload(file: UploadFile) -> List[Dict[str, Any]]:
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="请上传 .xlsx 或 .xlsm 文件")
    workbook = openpyxl.load_workbook(io.BytesIO(file.file.read()), read_only=True, data_only=True)
    return _read_result_rows_from_workbook(workbook, file.filename)


def load_records_from_path(path: Path) -> List[Dict[str, Any]]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    return _read_result_rows_from_workbook(workbook, path.name)


def summarize_records(records: List[Dict[str, Any]], top_n: int = 20) -> Dict[str, Any]:
    if not records:
        raise ValueError("没有可分析的分类记录")

    method_counter = Counter(record["method"] for record in records)
    level1_counter = Counter(record["level1"] for record in records)
    level2_counter = Counter(record["level2"] for record in records)
    structure_counter = Counter(record["structure_type"] for record in records)
    match_type_counter = Counter(record["match_type"] for record in records)

    focus_samples = [
        record
        for record in records
        if record["needs_review"] or record["is_composite"] or record["catalog_id"] == OUT_OF_SCOPE_ID
    ]
    focus_samples.sort(
        key=lambda item: (
            item["source_file"],
            not item["needs_review"],
            not item["is_composite"],
            item["row_num"],
        )
    )

    return {
        "summary": {
            "total_records": len(records),
            "rule_method_count": 0,
            "llm_method_count": method_counter.get("LLM主分类", 0),
            "fallback_method_count": method_counter.get("体系外默认分类", 0),
            "review_count": sum(1 for record in records if record["needs_review"]),
            "composite_count": sum(1 for record in records if record["is_composite"]),
        },
        "match_type_counts": {
            "single": match_type_counter.get("standard_catalog", 0),
            "cross_domain": 0,
            "same_domain_multi_item": 0,
            "low_confidence": 0,
            "llm_fallback": 0,
            "fallback": match_type_counter.get("out_of_scope", 0),
        },
        "structure_counts": {
            "single_project": structure_counter.get("single_project", 0),
            "composite_project": structure_counter.get("composite_project", 0),
            "multi_system_same_domain": 0,
        },
        "level1_top": [{"name": name, "count": count} for name, count in level1_counter.most_common(top_n)],
        "level2_top": [{"name": name, "count": count} for name, count in level2_counter.most_common(top_n)],
        "focus_samples": focus_samples[:FOCUS_SAMPLE_LIMIT],
    }


def analyze_excel_file(file: UploadFile, top_n: int = 20) -> Dict[str, Any]:
    records = load_records_from_upload(file)
    return summarize_records(records, top_n=top_n)
