import io

import openpyxl
from fastapi import HTTPException, UploadFile
from fastapi.responses import StreamingResponse

from services.standard_classifier import classify_project_standard


RESULT_HEADERS = [
    "工程名称",
    "catalog_id",
    "一级分类",
    "二级分类",
    "维修状态",
    "标准对象",
    "是否复合工程",
    "复合候选目录",
    "是否紧急维修",
    "是否白蚁相关",
    "是否建议复核",
    "候选目录",
    "分类依据",
]


def _bool_text(value: object) -> str:
    return "是" if bool(value) else "否"


def _write_result_row(worksheet, row: int, result: dict[str, object]) -> None:
    values = [
        result["project_name"],
        result["catalog_id"],
        result["category"],
        result["item"],
        result["repair_status"],
        result["standard_group"],
        _bool_text(result["is_composite"]),
        " | ".join(result["secondary_catalog_labels"]),
        _bool_text(result["is_emergency"]),
        _bool_text(result["termite_related"]),
        _bool_text(result["needs_review"]),
        " | ".join(result["candidate_labels"]),
        result["reason"],
    ]
    for column, value in enumerate(values, start=1):
        worksheet.cell(row=row, column=column, value=value)


def build_classified_workbook_bytes(data: bytes) -> bytes:
    source_workbook = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    source_sheet = source_workbook.active

    first_header = source_sheet.cell(row=1, column=1).value
    if first_header is None or str(first_header).strip() == "":
        raise HTTPException(status_code=400, detail="第一列表头不能为空")

    output_workbook = openpyxl.Workbook()
    output_sheet = output_workbook.active
    output_sheet.title = "分类结果"
    for column, header in enumerate(RESULT_HEADERS, start=1):
        output_sheet.cell(row=1, column=column, value=header)

    output_row = 2
    for source_row in range(2, source_sheet.max_row + 1):
        project_name = source_sheet.cell(row=source_row, column=1).value
        if project_name is None or str(project_name).strip() == "":
            continue
        result = classify_project_standard(str(project_name).strip())
        _write_result_row(output_sheet, output_row, result)
        output_row += 1

    output = io.BytesIO()
    output_workbook.save(output)
    return output.getvalue()


def classify_excel_file(file: UploadFile):
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="请上传 .xlsx 或 .xlsm 文件")

    output = io.BytesIO(build_classified_workbook_bytes(file.file.read()))
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
