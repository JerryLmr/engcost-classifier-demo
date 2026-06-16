import importlib.util
import os
import subprocess
import sys
import tempfile
import unittest
from io import BytesIO
from types import SimpleNamespace
from pathlib import Path
from unittest.mock import patch

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT / "backend"))

try:
    import openpyxl
except ImportError:  # pragma: no cover
    openpyxl = None

from api.routes import classify, health_check
from models.schemas import ClassifyRequest
from services.analysis_service import analyze_excel_file
from services.excel_service import build_classified_workbook_bytes


RESULT_HEADERS = [
    "一级分类",
    "二级分类",
    "三级分类",
    "具体细项",
    "分类方式",
    "置信度",
    "匹配类型",
    "是否建议复核",
    "候选目录ID",
    "候选目录",
    "候选细项",
    "分类依据",
]


def make_workbook(*project_names: str) -> bytes:
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.cell(row=1, column=1, value="工程名称")
    for row, value in enumerate(project_names, start=2):
        worksheet.cell(row=row, column=1, value=value)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


@unittest.skipIf(openpyxl is None, "测试依赖未安装")
class ApiAndExcelTestCase(unittest.TestCase):
    def test_health(self):
        data = health_check()
        self.assertEqual(data["status"], "ok")
        self.assertIn("model", data)

    def test_single_classify_response_shape(self):
        data = classify(ClassifyRequest(text="消防栓更换"))
        for field in [
            "level1",
            "level2",
            "level3_item",
            "matched_level3_items",
            "confidence",
            "match_type",
            "needs_review",
            "candidate_ids",
            "candidate_level3_items",
            "reason",
        ]:
            self.assertIn(field, data)
        self.assertNotIn("is_composite", data)
        self.assertEqual(data["level2"], "消防栓、箱")

    @patch("classifier.llm_client.request_llm_classification", side_effect=RuntimeError("offline"))
    def test_excel_classify_headers(self, _mock_request):
        output = build_classified_workbook_bytes(
            make_workbook("消防栓更换", "某小区综合整治提升项目")
        )
        workbook = openpyxl.load_workbook(BytesIO(output))
        worksheet = workbook.active
        headers = [worksheet.cell(row=1, column=i).value for i in range(1, worksheet.max_column + 1)]
        for header in RESULT_HEADERS:
            self.assertIn(header, headers)

    def test_analyze_excel_new_headers(self):
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        for col, value in enumerate(["工程名称", *RESULT_HEADERS], start=1):
            worksheet.cell(row=1, column=col, value=value)
        worksheet.append(
            [
                "消防栓更换",
                "消防工程",
                "消防栓、箱",
                "更换消防栓、箱",
                "更换消防栓、箱",
                "规则优先",
                "高",
                "single",
                "否",
                "038",
                "038 消防工程 > 消防栓、箱",
                "更换消防栓、箱",
                "命中对象词：消防栓",
            ]
        )
        output = BytesIO()
        workbook.save(output)

        data = analyze_excel_file(
            SimpleNamespace(
                filename="classified.xlsx",
                file=BytesIO(output.getvalue()),
            )
        )
        self.assertEqual(data["summary"]["total_records"], 1)
        self.assertEqual(data["match_type_counts"]["single"], 1)

    @patch("classifier.llm_client.request_llm_classification", side_effect=RuntimeError("offline"))
    def test_cli_single_file_and_directory_modes(self, _mock_request):
        script = ROOT / "scripts" / "batch_classify_excel.py"
        env = os.environ.copy()
        env["LLM_PROVIDER"] = "disabled"
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            input_file = tmp_path / "input.xlsx"
            input_file.write_bytes(make_workbook("消防栓更换"))
            output_file = tmp_path / "output.xlsx"

            single = subprocess.run(
                [sys.executable, str(script), str(input_file), "-o", str(output_file), "--overwrite"],
                cwd=ROOT,
                env=env,
                check=False,
                capture_output=True,
                text=True,
            )
            self.assertEqual(single.returncode, 0, single.stderr + single.stdout)
            self.assertTrue(output_file.exists())

            input_dir = tmp_path / "inputs"
            output_dir = tmp_path / "outputs"
            input_dir.mkdir()
            (input_dir / "batch.xlsx").write_bytes(make_workbook("屋面防水"))
            batch = subprocess.run(
                [sys.executable, str(script), str(input_dir), "-o", str(output_dir), "--overwrite"],
                cwd=ROOT,
                env=env,
                check=False,
                capture_output=True,
                text=True,
            )
            self.assertEqual(batch.returncode, 0, batch.stderr + batch.stdout)
            self.assertTrue((output_dir / "batch_classified.xlsx").exists())

    def test_cli_mode_dispatch(self):
        script = ROOT / "scripts" / "batch_classify_excel.py"
        spec = importlib.util.spec_from_file_location("batch_classify_excel", script)
        module = importlib.util.module_from_spec(spec)
        self.assertIsNotNone(spec)
        self.assertIsNotNone(spec.loader)
        spec.loader.exec_module(module)
        self.assertIn("catalog_id", module.RESULT_HEADERS)
        self.assertNotIn("置信度", module.RESULT_HEADERS)


if __name__ == "__main__":
    unittest.main()
