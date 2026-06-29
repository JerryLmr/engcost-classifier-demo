import importlib.util
import json
import os
import subprocess
import sys
import tempfile
import unittest
from io import BytesIO, StringIO
from contextlib import redirect_stdout
from pathlib import Path
from unittest.mock import patch

import openpyxl

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT / "backend"))

from classifier.alias_matcher import load_text_aliases, match_aliases
from classifier.catalog_postprocess import postprocess_item_selection
from classifier.llm_client import (
    ItemSelection,
    LLMServiceError,
    StatusSelection,
    build_full_catalog_item_selection_prompt,
    llm_select_catalog_item_from_full_catalog,
    llm_select_repair_status,
)
from classifier.standard_catalog_loader import (
    OUT_OF_SCOPE_ID,
    catalog_label,
    get_standard_catalog_by_id,
    load_standard_catalog,
)
from services.analysis_service import load_records_from_path
from services.standard_classifier import classify_project_standard


NEW_HEADERS = [
    "工程名称",
    "project_name_text",
    "catalog_id",
    "一级分类",
    "二级分类",
    "维修状态",
    "标准对象",
    "是否复合工程",
    "复合目录",
    "是否紧急维修",
    "是否白蚁相关",
    "是否建议复核",
    "分类依据",
    "file_name",
    "consultation_project_name",
    "renovation_content",
    "sub_project_id",
    "sub_item_project_rows",
    "consultation_time",
    "location",
    "cache_subject",
]

OLD_HEADERS = {"置信度", "匹配类型", "分类方式", "候选目录"}


def make_workbook(*project_names: str) -> bytes:
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    headers = [
        "file_name",
        "consultation_project_name",
        "consultation_time",
        "renovation_content",
        "sub_item_project_rows",
        "location",
    ]
    for column, header in enumerate(headers, start=1):
        worksheet.cell(row=1, column=column, value=header)
    for row, value in enumerate(project_names, start=2):
        worksheet.cell(row=row, column=1, value=f"file-{row}.pdf")
        worksheet.cell(row=row, column=2, value=value)
        worksheet.cell(row=row, column=3, value="2026-01-01")
        worksheet.cell(row=row, column=4, value="")
        worksheet.cell(row=row, column=5, value="[]")
        worksheet.cell(row=row, column=6, value="浙江省嘉兴市")
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def make_ocr_workbook(rows: list[tuple[object, object, object, object, object, object]]) -> bytes:
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    headers = [
        "file_name",
        "consultation_project_name",
        "consultation_time",
        "renovation_content",
        "sub_item_project_rows",
        "location",
    ]
    for column, header in enumerate(headers, start=1):
        worksheet.cell(row=1, column=column, value=header)
    for row_index, values in enumerate(rows, start=2):
        for column, value in enumerate(values, start=1):
            worksheet.cell(row=row_index, column=column, value=value)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def make_workbook_with_headers(headers: list[str], rows: list[list[object]]) -> bytes:
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    for column, header in enumerate(headers, start=1):
        worksheet.cell(row=1, column=column, value=header)
    for row_index, values in enumerate(rows, start=2):
        for column, value in enumerate(values, start=1):
            worksheet.cell(row=row_index, column=column, value=value)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def make_classified_workbook(path: Path) -> None:
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    for column, header in enumerate(NEW_HEADERS, start=1):
        worksheet.cell(row=1, column=column, value=header)
    values = [
        "屋面及外墙渗漏维修",
        "屋面外墙渗漏维修",
        "CP-002-03",
        "屋面",
        "防水层",
        "维修",
        "共用部位",
        "是",
        "CP-003-01 | 外墙面 | 面层",
        "否",
        "否",
        "是",
        "疑似复合工程",
        "ocr.xlsx",
        "咨询项目",
        "维修内容",
        "屋面及外墙渗漏维修",
        "[]",
        "2026-01-01",
        "浙江省嘉兴市",
        "屋面及外墙渗漏维修",
    ]
    for column, value in enumerate(values, start=1):
        worksheet.cell(row=2, column=column, value=value)
    workbook.save(path)


class StandardCatalogPipelineTestCase(unittest.TestCase):
    def test_standard_catalog_loads_required_items(self):
        catalog = load_standard_catalog()
        catalog_by_id = get_standard_catalog_by_id()
        self.assertGreater(len(catalog), 0)
        self.assertNotIn(OUT_OF_SCOPE_ID, catalog_by_id)
        self.assertEqual(len(catalog_by_id), len(catalog))
        for item_id in ("CF-017-00", "CF-018-00", "CF-028-00", "TERMITE-001"):
            self.assertIn(item_id, catalog_by_id)
        for item in catalog:
            self.assertTrue(item.status_basis)
            self.assertEqual(catalog_label(item), f"{item.id} | {item.category} | {item.item}")

    def test_text_aliases_load_and_only_expand_terms(self):
        entries = load_text_aliases()
        self.assertGreater(len(entries), 0)
        result = match_aliases("致远大厦外立面修缮工程")
        self.assertIn("外墙面", result.expanded_terms)
        self.assertFalse(hasattr(result, "catalog_hits"))
        self.assertFalse(hasattr(result, "negative_hints"))
        self.assertFalse(hasattr(result, "review_hints"))

    def test_full_catalog_prompt_uses_compact_lines_without_status_basis(self):
        catalog = [get_standard_catalog_by_id()["CF-015-05"]]
        prompt = build_full_catalog_item_selection_prompt(
            "报审项目名称",
            "减压阀更换",
            catalog,
            ["阀门"],
            ["alias辅助扩展词（不能直接决定分类）：阀门"],
        )
        self.assertIn("catalog_id | 标准对象 | 一级分类 | 二级分类 | 目录属性 | 可选状态", prompt)
        self.assertIn("CF-015-05 | 共用设施设备 | 给排水系统", prompt)
        self.assertIn("标准项", prompt)
        self.assertIn("报审项目名称：报审项目名称", prompt)
        self.assertIn("本次分类对象：减压阀更换", prompt)
        self.assertIn("清单摘要", prompt)
        self.assertIn("1. 阀门", prompt)
        self.assertIn("alias辅助扩展词", prompt)
        self.assertIn("project_name_text", prompt)
        self.assertIn("用于相似项目检索的工程语义文本", prompt)
        self.assertNotIn("状态依据", prompt)
        self.assertNotIn("候选目录：", prompt)
        self.assertNotIn("renovation_content", prompt)

    def test_full_catalog_prompt_marks_unspecified_items_as_fallback_and_sorts_them_last(self):
        catalog_by_id = get_standard_catalog_by_id()
        catalog = [
            catalog_by_id["CF-018-00"],
            catalog_by_id["CF-018-02"],
            catalog_by_id["CF-018-09"],
            catalog_by_id["CF-018-10"],
            catalog_by_id["CF-018-12"],
            catalog_by_id["CF-015-05"],
        ]
        prompt = build_full_catalog_item_selection_prompt(
            "平湖市碧水云天监控改造",
            "平湖市碧水云天监控改造-安装",
            catalog,
            ["摄像头", "网络录像机", "交换机", "监控硬盘", "光纤收发器", "电梯网桥"],
        )

        self.assertIn("catalog_id | 标准对象 | 一级分类 | 二级分类 | 目录属性 | 可选状态", prompt)
        self.assertIn("CF-018-00 | 共用设施设备 | 弱电系统 | 未明确具体子项 | 内部扩展项", prompt)
        self.assertIn("CF-015-05 | 共用设施设备 | 给排水系统", prompt)
        self.assertIn("标准项", prompt)
        self.assertIn("最具体可匹配原则", prompt)
        self.assertIn("兜底项，不是泛化首选项", prompt)
        self.assertIn("只有在无法从“本次分类对象”和“清单摘要”判断任何具体二级对象时", prompt)
        self.assertNotIn("如果只能判断一级系统，且标准目录里存在“未明确具体子项”，可以选择该项", prompt)
        self.assertIn("本次分类对象：平湖市碧水云天监控改造-安装", prompt)
        self.assertNotIn("cache_subject", prompt)

        unspecified_index = prompt.index("CF-018-00 | 共用设施设备 | 弱电系统 | 未明确具体子项")
        for item_id in ("CF-018-02", "CF-018-09", "CF-018-10", "CF-018-12"):
            self.assertLess(prompt.index(f"{item_id} | 共用设施设备 | 弱电系统"), unspecified_index)

    @patch(
        "classifier.llm_client.request_llm_json",
        return_value={
            "catalog_id": "CF-015-05",
            "secondary_catalog_ids": ["BAD-ID", "CF-015-04"],
            "is_composite": True,
            "needs_review": False,
            "reason": "对象为减压阀和管道",
            "project_name_text": "减压阀管道维修",
        },
    )
    def test_full_catalog_selection_filters_invalid_secondary_ids(self, _mock_request):
        result = llm_select_catalog_item_from_full_catalog("报审项目", "减压阀及管道维修", load_standard_catalog())
        self.assertEqual(result.catalog_id, "CF-015-05")
        self.assertEqual(result.secondary_catalog_ids, ("CF-015-04",))
        self.assertEqual(result.project_name_text, "减压阀管道维修")
        self.assertTrue(result.needs_review)
        self.assertIn("已丢弃标准外 secondary id", result.reason)

    @patch(
        "classifier.llm_client.request_llm_json",
        return_value={
            "catalog_id": "BAD-ID",
            "secondary_catalog_ids": [],
            "is_composite": False,
            "needs_review": False,
            "reason": "bad",
        },
    )
    def test_full_catalog_selection_falls_back_after_invalid_primary_id(self, mock_request):
        result = llm_select_catalog_item_from_full_catalog("报审项目", "未知项目", load_standard_catalog())
        self.assertEqual(result.catalog_id, OUT_OF_SCOPE_ID)
        self.assertTrue(result.needs_review)
        self.assertTrue(result.invalid_after_retry)
        self.assertIn("LLM returned invalid catalog_id: BAD-ID", result.reason)
        self.assertEqual(mock_request.call_count, 2)

    @patch("classifier.llm_client.request_llm_json", side_effect=LLMServiceError("service down"))
    def test_full_catalog_selection_raises_service_error_without_retry(self, mock_request):
        with self.assertRaises(LLMServiceError):
            llm_select_catalog_item_from_full_catalog("报审项目", "减压阀更换", load_standard_catalog())
        self.assertEqual(mock_request.call_count, 1)

    def test_postprocess_promotes_weak_current_secondary_item(self):
        result = postprocess_item_selection(
            "弱电安防、监控、门禁",
            ItemSelection(
                catalog_id="CF-018-00",
                secondary_catalog_ids=("CF-018-02", "CF-018-03"),
                is_composite=False,
                needs_review=False,
                reason="测试固定目录",
            ),
            get_standard_catalog_by_id(),
        )
        self.assertEqual(result.catalog_id, "CF-018-02")
        self.assertEqual(result.secondary_catalog_ids, ("CF-018-03",))
        self.assertTrue(result.needs_review)
        self.assertTrue(result.is_composite)
        self.assertIn("已将具体子项提升为主分类", result.reason)

    def test_postprocess_promotes_elevator_secondary_item(self):
        result = postprocess_item_selection(
            "电梯钢丝绳和曳引轮",
            ItemSelection(
                catalog_id="CF-017-00",
                secondary_catalog_ids=("CF-017-05", "CF-017-04"),
                is_composite=False,
                needs_review=False,
                reason="测试固定目录",
            ),
            get_standard_catalog_by_id(),
        )
        self.assertEqual(result.catalog_id, "CF-017-05")
        self.assertEqual(result.secondary_catalog_ids, ("CF-017-04",))
        self.assertTrue(result.needs_review)
        self.assertTrue(result.is_composite)

    def test_postprocess_keeps_unspecified_item_without_secondary(self):
        selection = ItemSelection(
            catalog_id="CF-018-00",
            secondary_catalog_ids=(),
            is_composite=False,
            needs_review=True,
            reason="仅明确弱电系统",
        )
        result = postprocess_item_selection("弱电系统维修", selection, get_standard_catalog_by_id())
        self.assertEqual(result, selection)

    def test_postprocess_keeps_specific_primary_item(self):
        selection = ItemSelection(
            catalog_id="CF-018-02",
            secondary_catalog_ids=("CF-018-03",),
            is_composite=True,
            needs_review=True,
            reason="测试固定目录",
        )
        result = postprocess_item_selection("监控和门禁", selection, get_standard_catalog_by_id())
        self.assertEqual(result, selection)

    def test_roof_leak_postprocess_prefers_waterproof_layer(self):
        result = postprocess_item_selection(
            "屋面渗漏水维修",
            ItemSelection(
                catalog_id="CP-002-01",
                secondary_catalog_ids=(),
                is_composite=False,
                needs_review=False,
                reason="测试固定目录",
            ),
            get_standard_catalog_by_id(),
        )
        self.assertEqual(result.catalog_id, "CP-002-03")
        self.assertEqual(result.secondary_catalog_ids, ())
        self.assertTrue(result.needs_review)
        self.assertFalse(result.is_composite)
        self.assertIn("屋面渗漏水优先归入防水层", result.reason)

    def test_roof_leak_postprocess_adds_exterior_wall_secondary(self):
        result = postprocess_item_selection(
            "屋面防水补漏及外墙渗漏",
            ItemSelection(
                catalog_id="CP-002-01",
                secondary_catalog_ids=(),
                is_composite=False,
                needs_review=False,
                reason="测试固定目录",
            ),
            get_standard_catalog_by_id(),
        )
        self.assertEqual(result.catalog_id, "CP-002-03")
        self.assertEqual(result.secondary_catalog_ids, ("CP-003-01",))
        self.assertTrue(result.needs_review)
        self.assertTrue(result.is_composite)

    def test_roof_leak_postprocess_keeps_roof_structure_when_structure_terms_exist(self):
        selection = ItemSelection(
            catalog_id="CP-002-01",
            secondary_catalog_ids=(),
            is_composite=False,
            needs_review=False,
            reason="测试固定目录",
        )
        result = postprocess_item_selection("屋脊瓦面脱落维修", selection, get_standard_catalog_by_id())
        self.assertEqual(result, selection)

    def test_roof_leak_postprocess_ignores_non_roof_structure_primary(self):
        selection = ItemSelection(
            catalog_id="CP-002-03",
            secondary_catalog_ids=(),
            is_composite=False,
            needs_review=False,
            reason="测试固定目录",
        )
        result = postprocess_item_selection("屋面渗漏水维修", selection, get_standard_catalog_by_id())
        self.assertEqual(result, selection)

    @patch(
        "classifier.llm_client.request_llm_json",
        return_value={
            "repair_status": "非法状态",
            "needs_review": False,
            "reason": "bad",
        },
    )
    def test_llm_status_selection_falls_back_after_invalid_status_retry(self, mock_request):
        item = get_standard_catalog_by_id()["CF-015-05"]
        result = llm_select_repair_status("减压阀更换", item)
        self.assertEqual(result.repair_status, "不确定")
        self.assertTrue(result.needs_review)
        self.assertTrue(result.invalid_after_retry)
        self.assertEqual(mock_request.call_count, 2)

    @patch("classifier.llm_client.request_llm_json", side_effect=LLMServiceError("service down"))
    def test_llm_status_selection_raises_service_error_without_retry(self, mock_request):
        item = get_standard_catalog_by_id()["CF-015-05"]
        with self.assertRaises(LLMServiceError):
            llm_select_repair_status("减压阀更换", item)
        self.assertEqual(mock_request.call_count, 1)

    @patch("classifier.llm_client._request_lmstudio_json", return_value={"ok": True})
    def test_request_llm_json_uses_lmstudio_directly(self, mock_request):
        from classifier.llm_client import request_llm_json

        self.assertEqual(request_llm_json("prompt"), {"ok": True})
        mock_request.assert_called_once_with("prompt", max_tokens=None, timeout_seconds=None, system_prompt=None)

    @patch(
        "classifier.llm_client.llm_select_repair_status",
        return_value=StatusSelection(repair_status="更新", needs_review=False, reason="更换对应更新"),
    )
    @patch(
        "classifier.llm_client.llm_select_catalog_item_from_full_catalog",
        return_value=ItemSelection(
            catalog_id="CF-015-05",
            secondary_catalog_ids=(),
            is_composite=False,
            needs_review=False,
            reason="对象为减压阀",
            project_name_text="减压阀更换",
        ),
    )
    def test_standard_classifier_result_shape_has_no_candidate_labels(self, _mock_item, _mock_status):
        result = classify_project_standard("减压阀更换")
        self.assertEqual(result["catalog_id"], "CF-015-05")
        self.assertEqual(result["repair_status"], "更新")
        self.assertEqual(result["project_name_text"], "减压阀更换")
        self.assertNotIn("candidate_labels", result)
        self.assertEqual(result["secondary_catalog_ids"], [])
        self.assertEqual(result["secondary_catalog_labels"], [])
        self.assertFalse(result["needs_review"])

    @patch(
        "classifier.llm_client.llm_select_catalog_item_from_full_catalog",
        side_effect=LLMServiceError("service down"),
    )
    def test_standard_classifier_marks_llm_service_error(self, _mock_item):
        result = classify_project_standard("减压阀更换")
        self.assertEqual(result["pipeline_status"], "llm_service_error")
        self.assertEqual(result["catalog_id"], OUT_OF_SCOPE_ID)
        self.assertIn("service down", result["reason"])

    @patch(
        "classifier.llm_client.llm_select_repair_status",
        return_value=StatusSelection(repair_status="维修", needs_review=False, reason="按工程名称判断为维修"),
    )
    @patch("classifier.llm_client.llm_select_catalog_item_from_full_catalog")
    def test_full_catalog_regression_cases_use_llm_selection_without_candidates(self, mock_item, _mock_status):
        expected_by_text = {
            "篮球场和网球场改造工程 篮球场和网球场地改造、更新": ("CF-011-02", (), True),
            "楼道粉刷": ("CP-004-02", (), False),
            "楼道整新": ("CP-004-02", (), False),
            "楼道整改": ("CP-004-02", (), False),
            "安防、人脸识别、监控、门禁": ("CF-018-02", ("CF-018-03",), True),
            "车牌识别系统更新": ("CF-018-07", (), False),
            "君临天下花园人行道维修": ("CF-007-02", (), False),
            "山鑫康城二次供水维修更新 二次供水改直供水": ("CF-015-02", ("CF-015-06",), True),
            "岛语树雅苑电梯钢带维修工程 电梯钢带维修": ("CF-017-05", (), False),
            "电梯维修": ("CF-017-00", (), True),
            "消防改造": ("CF-028-00", (), True),
            "弱电智能化改造": ("CF-018-00", (), True),
            "电梯监控维修": ("CF-018-02", (), False),
            "轿厢监控系统更新": ("CF-018-02", (), False),
            "消防检测维修": ("CF-028-00", (), True),
            "电梯维保维修": ("CF-017-00", (), True),
            "美丽家园改造：道路、绿化、监控": ("CF-007-02", ("CF-018-02",), True),
        }

        def select_item(_consultation_project_name, classify_subject, _catalog, _item_summary=None, _context_hints=None):
            catalog_id, secondary_ids, needs_review = expected_by_text[classify_subject]
            return ItemSelection(
                catalog_id=catalog_id,
                secondary_catalog_ids=secondary_ids,
                is_composite=bool(secondary_ids),
                needs_review=needs_review,
                reason="测试固定目录",
            )

        mock_item.side_effect = select_item

        for text, (expected_id, secondary_ids, expected_review) in expected_by_text.items():
            with self.subTest(text=text):
                result = classify_project_standard(text)
                self.assertEqual(result["catalog_id"], expected_id)
                self.assertNotEqual(result["catalog_id"], OUT_OF_SCOPE_ID)
                self.assertEqual(result["secondary_catalog_ids"], list(secondary_ids))
                self.assertNotIn("candidate_labels", result)
                if expected_review or secondary_ids:
                    self.assertTrue(result["needs_review"])

    @patch(
        "classifier.llm_client.llm_select_catalog_item_from_full_catalog",
        side_effect=[
            ItemSelection(
                catalog_id=OUT_OF_SCOPE_ID,
                secondary_catalog_ids=(),
                is_composite=False,
                needs_review=True,
                reason="宏观项目未明确具体维修对象",
            ),
            ItemSelection(
                catalog_id=OUT_OF_SCOPE_ID,
                secondary_catalog_ids=(),
                is_composite=False,
                needs_review=True,
                reason="综合整治未明确具体维修对象",
            ),
        ],
    )
    def test_macro_projects_can_stay_out_of_scope_without_candidate_fallback(self, _mock_item):
        for text in ("美丽家园改造", "小区综合整治工程"):
            with self.subTest(text=text):
                result = classify_project_standard(text)
                self.assertEqual(result["catalog_id"], OUT_OF_SCOPE_ID)
                self.assertTrue(result["needs_review"])
                self.assertNotIn("candidate_labels", result)

    @patch(
        "classifier.llm_client.llm_select_repair_status",
        return_value=StatusSelection(repair_status="维修", needs_review=False, reason="按工程名称判断为维修"),
    )
    @patch(
        "classifier.llm_client.llm_select_catalog_item_from_full_catalog",
        return_value=ItemSelection(
            catalog_id="CP-002-03",
            secondary_catalog_ids=("CP-003-01",),
            is_composite=True,
            needs_review=True,
            reason="屋面及外墙复合工程",
        ),
    )
    def test_secondary_catalog_labels_are_preserved(self, _mock_item, _mock_status):
        result = classify_project_standard("屋面及外墙渗漏维修")
        self.assertEqual(result["secondary_catalog_ids"], ["CP-003-01"])
        self.assertEqual(result["secondary_catalog_labels"], [catalog_label(get_standard_catalog_by_id()["CP-003-01"])])
        self.assertNotIn("candidate_labels", result)

    def test_analysis_service_accepts_outputs_without_candidate_column(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "classified.xlsx"
            make_classified_workbook(path)
            records = load_records_from_path(path)
        self.assertEqual(len(records), 1)
        self.assertNotIn("candidate_labels", records[0])
        self.assertEqual(records[0]["secondary_candidates"], ["CP-003-01", "外墙面", "面层"])

    def test_batch_script_outputs_new_headers_only(self):
        script = ROOT / "scripts" / "batch_classify_excel.py"
        spec = importlib.util.spec_from_file_location("batch_classify_excel", script)
        module = importlib.util.module_from_spec(spec)
        self.assertIsNotNone(spec)
        self.assertIsNotNone(spec.loader)
        spec.loader.exec_module(module)

        def fake_classify(project_text: str, **_kwargs) -> dict[str, object]:
            return {
                "project_name": project_text,
                "catalog_id": OUT_OF_SCOPE_ID,
                "category": "体系外",
                "item": "体系外",
                "pipeline_status": "fallback",
            }

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            input_file = tmp_path / "input.xlsx"
            output_file = tmp_path / "output.xlsx"
            input_file.write_bytes(make_workbook("减压阀更换", "屋面及外墙渗漏维修"))
            module.classify_workbook(input_file, output_file, fake_classify, {})
            workbook = openpyxl.load_workbook(output_file)
            worksheet = workbook.active
            headers = [worksheet.cell(row=1, column=i).value for i in range(1, worksheet.max_column + 1)]
            self.assertEqual(headers, NEW_HEADERS)
            self.assertTrue(OLD_HEADERS.isdisjoint(set(headers)))
            self.assertEqual(headers[-1], "cache_subject")
            self.assertNotIn("cache_key", headers)
            self.assertNotIn("semantic_cache_key", headers)
            self.assertEqual(worksheet.cell(row=2, column=1).value, "减压阀更换")
            self.assertEqual(worksheet.cell(row=2, column=2).value, "减压阀更换")
            self.assertEqual(worksheet.cell(row=2, column=3).value, OUT_OF_SCOPE_ID)

    def test_batch_script_preflight_failure_returns_error_without_output(self):
        script = ROOT / "scripts" / "batch_classify_excel.py"
        env = os.environ.copy()
        env["LMSTUDIO_BASE_URL"] = "http://127.0.0.1:9/v1"
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            input_file = tmp_path / "input.xlsx"
            output_file = tmp_path / "output.xlsx"
            input_file.write_bytes(make_workbook("减压阀更换"))
            run = subprocess.run(
                [sys.executable, str(script), str(input_file), "-o", str(output_file), "--overwrite"],
                cwd=ROOT,
                env=env,
                check=False,
                capture_output=True,
                text=True,
                timeout=10,
            )
            self.assertEqual(run.returncode, 1)
            self.assertIn("[ERROR] LM Studio 服务不可用", run.stdout)
            self.assertFalse(output_file.exists())

    def test_batch_script_llm_service_error_fails_without_writing_output(self):
        script = ROOT / "scripts" / "batch_classify_excel.py"
        spec = importlib.util.spec_from_file_location("batch_classify_excel", script)
        module = importlib.util.module_from_spec(spec)
        self.assertIsNotNone(spec)
        self.assertIsNotNone(spec.loader)
        spec.loader.exec_module(module)

        def fake_classify(project_text: str, **_kwargs) -> dict[str, object]:
            return {
                "project_name": project_text,
                "catalog_id": OUT_OF_SCOPE_ID,
                "pipeline_status": "llm_service_error",
                "reason": "service down",
            }

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            input_file = tmp_path / "input.xlsx"
            output_file = tmp_path / "output.xlsx"
            input_file.write_bytes(make_workbook("减压阀更换"))
            with self.assertRaisesRegex(RuntimeError, "LLM 服务连接失败"):
                module.classify_workbook(input_file, output_file, fake_classify, {})
            self.assertFalse(output_file.exists())

    def test_batch_script_caches_repeated_project_names(self):
        script = ROOT / "scripts" / "batch_classify_excel.py"
        spec = importlib.util.spec_from_file_location("batch_classify_excel", script)
        module = importlib.util.module_from_spec(spec)
        self.assertIsNotNone(spec)
        self.assertIsNotNone(spec.loader)
        spec.loader.exec_module(module)

        calls: list[str] = []

        def fake_classify(project_text: str, **_kwargs) -> dict[str, object]:
            calls.append(project_text)
            return {
                "project_name": project_text,
                "catalog_id": f"ID-{len(calls)}",
                "category": "一级",
                "item": "二级",
                "pipeline_status": "ok",
            }

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            input_file = tmp_path / "input.xlsx"
            output_file = tmp_path / "output.xlsx"
            input_file.write_bytes(make_workbook("屋面维修", "屋面维修", "", "外墙维修"))

            cache: dict[str, dict[str, object]] = {}
            stats = module.classify_workbook(input_file, output_file, fake_classify, cache)

            self.assertEqual(stats, (3, 1, 4, 1))
            self.assertEqual(calls, ["屋面维修", "未分组", "外墙维修"])
            self.assertFalse(any("屋面维修" in key for key in cache))
            self.assertNotIn("屋面维修", cache)
            workbook = openpyxl.load_workbook(output_file)
            worksheet = workbook.active
            self.assertEqual(worksheet.cell(row=2, column=2).value, "屋面维修")
            self.assertEqual(worksheet.cell(row=2, column=3).value, "ID-1")
            self.assertEqual(worksheet.cell(row=3, column=3).value, "ID-1")
            self.assertEqual(worksheet.cell(row=4, column=1).value, "未分组")
            self.assertEqual(worksheet.cell(row=4, column=3).value, "ID-2")
            self.assertEqual(worksheet.cell(row=5, column=3).value, "ID-3")

    def test_batch_script_cache_reuses_across_files_and_preserves_ocr_fields(self):
        script = ROOT / "scripts" / "batch_classify_excel.py"
        spec = importlib.util.spec_from_file_location("batch_classify_excel", script)
        module = importlib.util.module_from_spec(spec)
        self.assertIsNotNone(spec)
        self.assertIsNotNone(spec.loader)
        spec.loader.exec_module(module)

        calls: list[str] = []

        def fake_classify(project_text: str, **_kwargs) -> dict[str, object]:
            calls.append(project_text)
            return {
                "project_name": project_text,
                "catalog_id": "CP-TEST",
                "category": "一级",
                "item": "二级",
                "pipeline_status": "ok",
            }

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            first_input = tmp_path / "first.xlsx"
            second_input = tmp_path / "second.xlsx"
            first_output = tmp_path / "first-output.xlsx"
            second_output = tmp_path / "second-output.xlsx"
            first_input.write_bytes(
                make_ocr_workbook(
                    [
                        ("file-a.pdf", "小区屋面", "2026-01-01", "防水维修", "[{\"seq\": 1}]", "浙江省嘉兴市"),
                        ("file-b.pdf", "小区屋面", "2026-01-01", "防水维修", "[{\"seq\": 2}]", "浙江省嘉兴市"),
                        ("blank.pdf", "", "2026-01-01", "", "[{\"seq\": 3}]", "浙江省嘉兴市"),
                    ]
                )
            )
            second_input.write_bytes(
                make_ocr_workbook([("file-c.pdf", "小区屋面", "2026-01-01", "防水维修", "[]", "浙江省嘉兴市")])
            )

            cache: dict[str, dict[str, object]] = {}
            first_stats = module.classify_workbook(first_input, first_output, fake_classify, cache)
            second_stats = module.classify_workbook(second_input, second_output, fake_classify, cache)

            self.assertEqual(first_stats, (2, 1, 3, 1))
            self.assertEqual(second_stats, (0, 1, 1, 0))
            self.assertEqual(calls, ["小区屋面", "未分组"])

            first_workbook = openpyxl.load_workbook(first_output)
            first_sheet = first_workbook.active
            self.assertEqual(first_sheet.cell(row=2, column=14).value, "file-a.pdf")
            self.assertEqual(first_sheet.cell(row=2, column=17).value, "小区屋面")
            self.assertIn('"seq": 1', first_sheet.cell(row=2, column=18).value)
            self.assertEqual(first_sheet.cell(row=3, column=14).value, "file-b.pdf")
            self.assertEqual(first_sheet.cell(row=3, column=17).value, "小区屋面")
            self.assertIn('"seq": 2', first_sheet.cell(row=3, column=18).value)
            self.assertEqual(first_sheet.cell(row=4, column=14).value, "blank.pdf")
            self.assertEqual(first_sheet.cell(row=4, column=17).value, "未分组")
            self.assertIn('"seq": 3', first_sheet.cell(row=4, column=18).value)

            second_workbook = openpyxl.load_workbook(second_output)
            second_sheet = second_workbook.active
            self.assertEqual(second_sheet.cell(row=2, column=1).value, "小区屋面")
            self.assertEqual(second_sheet.cell(row=2, column=14).value, "file-c.pdf")

    def test_batch_script_unit_project_name_and_cache_subject_for_numbered_rooms(self):
        script = ROOT / "scripts" / "batch_classify_excel.py"
        spec = importlib.util.spec_from_file_location("batch_classify_excel", script)
        module = importlib.util.module_from_spec(spec)
        self.assertIsNotNone(spec)
        self.assertIsNotNone(spec.loader)
        spec.loader.exec_module(module)

        calls: list[str] = []

        def fake_classify(project_text: str, **kwargs) -> dict[str, object]:
            calls.append(project_text)
            self.assertEqual(project_text, "A小区屋面维修工程-15幢-1101")
            self.assertEqual(kwargs["consultation_project_name"], "A小区屋面维修工程")
            return {
                "project_name": project_text,
                "project_name_text": "屋面维修工程",
                "catalog_id": "CP-002-03",
                "category": "屋面",
                "item": "防水层",
                "pipeline_status": "ok",
            }

        rows = [
            {
                "page_no": 1,
                "seq": 1,
                "sub_project_id": "A小区屋面维修工程-15幢-1101",
                "project_code": "0101",
                "project_name": "屋面防水",
                "total_price": 100,
            },
            {
                "page_no": 1,
                "seq": 2,
                "sub_project_id": "A小区屋面维修工程-15幢-1102",
                "project_code": "0102",
                "project_name": "屋面防水",
                "total_price": 200,
            },
        ]

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            input_file = tmp_path / "input.xlsx"
            output_file = tmp_path / "output.xlsx"
            input_file.write_bytes(
                make_ocr_workbook(
                    [
                        (
                            "roof.pdf",
                            "A小区屋面维修工程",
                            "2026-01-01",
                            "屋面维修",
                            json.dumps(rows, ensure_ascii=False),
                            "浙江省嘉兴市",
                        )
                    ]
                )
            )

            cache: dict[str, dict[str, object]] = {}
            stats = module.classify_workbook(input_file, output_file, fake_classify, cache)

            self.assertEqual(stats, (1, 1, 2, 0))
            self.assertEqual(calls, ["A小区屋面维修工程-15幢-1101"])
            self.assertEqual(len(cache), 1)
            self.assertEqual(
                module._classification_cache_key("屋面维修", "A小区屋面维修工程-15幢-1101"),
                module._classification_cache_key("屋面维修", "A小区屋面维修工程-15幢-1102"),
            )

            workbook = openpyxl.load_workbook(output_file)
            worksheet = workbook.active
            headers = [worksheet.cell(row=1, column=i).value for i in range(1, worksheet.max_column + 1)]
            self.assertNotIn("cache_key", headers)
            self.assertNotIn("semantic_cache_key", headers)
            self.assertEqual(worksheet.cell(row=2, column=1).value, "A小区屋面维修工程-15幢-1101")
            self.assertEqual(worksheet.cell(row=3, column=1).value, "A小区屋面维修工程-15幢-1102")
            self.assertEqual(worksheet.cell(row=2, column=17).value, "A小区屋面维修工程-15幢-1101")
            self.assertEqual(worksheet.cell(row=3, column=17).value, "A小区屋面维修工程-15幢-1102")
            self.assertEqual(worksheet.cell(row=2, column=21).value, "A小区屋面维修工程")
            self.assertEqual(worksheet.cell(row=3, column=21).value, "A小区屋面维修工程")

    def test_batch_script_display_name_prefixes_consultation_when_needed(self):
        script = ROOT / "scripts" / "batch_classify_excel.py"
        spec = importlib.util.spec_from_file_location("batch_classify_excel", script)
        module = importlib.util.module_from_spec(spec)
        self.assertIsNotNone(spec)
        self.assertIsNotNone(spec.loader)
        spec.loader.exec_module(module)

        def fake_classify(project_text: str, **_kwargs) -> dict[str, object]:
            self.assertEqual(project_text, "15幢-1101")
            return {
                "project_name": project_text,
                "project_name_text": "15幢-1101",
                "catalog_id": "CP-002-03",
                "category": "屋面",
                "item": "防水层",
                "pipeline_status": "ok",
            }

        rows = [
            {
                "page_no": 1,
                "seq": 1,
                "sub_project_id": "15幢-1101",
                "project_code": "0101",
                "project_name": "屋面防水",
                "total_price": 100,
            }
        ]

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            input_file = tmp_path / "input.xlsx"
            output_file = tmp_path / "output.xlsx"
            input_file.write_bytes(
                make_ocr_workbook(
                    [
                        (
                            "room.pdf",
                            "A小区屋面维修工程",
                            "2026-01-01",
                            "屋面维修",
                            json.dumps(rows, ensure_ascii=False),
                            "浙江省嘉兴市",
                        )
                    ]
                )
            )

            module.classify_workbook(input_file, output_file, fake_classify, {})

            workbook = openpyxl.load_workbook(output_file)
            worksheet = workbook.active
            self.assertEqual(worksheet.cell(row=2, column=1).value, "A小区屋面维修工程-15幢-1101")
            self.assertEqual(worksheet.cell(row=2, column=17).value, "15幢-1101")
            self.assertEqual(worksheet.cell(row=2, column=21).value, "A小区屋面维修工程")

    def test_batch_script_uses_unit_project_name_for_cache_and_logs_only(self):
        script = ROOT / "scripts" / "batch_classify_excel.py"
        spec = importlib.util.spec_from_file_location("batch_classify_excel", script)
        module = importlib.util.module_from_spec(spec)
        self.assertIsNotNone(spec)
        self.assertIsNotNone(spec.loader)
        spec.loader.exec_module(module)

        consultation_name = "平湖滨水广场消火栓主管漏水，泵房压力下降及更换消防报警主机项目"
        unit_project_name = f"{consultation_name}-单项工程-安装"
        calls: list[str] = []

        def fake_classify(project_text: str, **_kwargs) -> dict[str, object]:
            calls.append(project_text)
            return {
                "project_name": project_text,
                "project_name_text": "消防报警主机安装",
                "catalog_id": "CF-028-00",
                "category": "消防系统",
                "item": "未明确具体子项",
                "pipeline_status": "ok",
            }

        rows = [
            {
                "page_no": 1,
                "seq": 1,
                "sub_project_id": "单项工程-安装",
                "project_code": "030404035001",
                "project_name": "消防报警主机",
                "total_price": 100,
            }
        ]

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            input_file = tmp_path / "input.xlsx"
            output_file = tmp_path / "output.xlsx"
            input_file.write_bytes(
                make_ocr_workbook(
                    [
                        (
                            "fire.pdf",
                            consultation_name,
                            "2026-01-01",
                            "安装工程",
                            json.dumps(rows, ensure_ascii=False),
                            "浙江省嘉兴市",
                        )
                    ]
                )
            )

            text_stdout = StringIO()
            with redirect_stdout(text_stdout):
                module.classify_workbook(input_file, output_file, fake_classify, {})

            self.assertEqual(calls, ["单项工程-安装"])
            log_text = text_stdout.getvalue()
            self.assertIn(unit_project_name, log_text)
            self.assertIn(f"cache_subject={unit_project_name}", log_text)
            self.assertNotIn("[ROW ] input.xlsx:2 单项工程-安装 cache_subject=单项工程-安装", log_text)

            workbook = openpyxl.load_workbook(output_file)
            worksheet = workbook.active
            self.assertEqual(worksheet.cell(row=2, column=1).value, unit_project_name)
            self.assertEqual(worksheet.cell(row=2, column=17).value, "单项工程-安装")
            self.assertEqual(worksheet.cell(row=2, column=21).value, unit_project_name)
            self.assertNotEqual(worksheet.cell(row=2, column=21).value, "单项工程-安装")

    def test_batch_script_cache_subject_normalizes_position_numbers_conservatively(self):
        script = ROOT / "scripts" / "batch_classify_excel.py"
        spec = importlib.util.spec_from_file_location("batch_classify_excel", script)
        module = importlib.util.module_from_spec(spec)
        self.assertIsNotNone(spec)
        self.assertIsNotNone(spec.loader)
        spec.loader.exec_module(module)

        self.assertFalse(hasattr(module, "_display_project_name"))
        self.assertEqual(
            module._classification_cache_key.__code__.co_varnames[:2],
            ("renovation_content", "unit_project_name"),
        )
        self.assertEqual(
            module._classification_cache_subject("锦绣庄园屋面维修工程-15幢-1101"),
            "锦绣庄园屋面维修工程",
        )
        self.assertEqual(
            module._classification_cache_subject("锦绣庄园屋面维修工程-15幢-1001、1002"),
            "锦绣庄园屋面维修工程",
        )
        self.assertEqual(
            module._classification_cache_subject("锦绣庄园屋面维修工程-15幢-1001,1002室"),
            "锦绣庄园屋面维修工程",
        )
        self.assertEqual(
            module._classification_cache_subject("锦绣庄园屋面维修工程-15幢-1001，1002室"),
            "锦绣庄园屋面维修工程",
        )
        self.assertEqual(module._classification_cache_subject("15幢屋面维修"), "屋面维修")
        self.assertEqual(module._classification_cache_subject("16幢屋面维修"), "屋面维修")
        self.assertNotEqual(
            module._classification_cache_subject("15幢屋面维修"),
            module._classification_cache_subject("16幢外墙维修"),
        )
        self.assertEqual(
            module._classification_cache_key("屋面维修", "锦绣庄园屋面维修工程-15幢-1101"),
            module._classification_cache_key("屋面维修", "锦绣庄园屋面维修工程-15幢-1001、1002"),
        )
        self.assertNotEqual(
            module._classification_cache_key("屋面维修", "15幢屋面维修"),
            module._classification_cache_key("屋面维修", "16幢外墙维修"),
        )
        self.assertEqual(
            module._build_unit_project_name(
                "锦绣庄园屋面维修工程15幢",
                "锦绣庄园屋面维修工程-15幢-1101",
            ),
            "锦绣庄园屋面维修工程-15幢-1101",
        )
        self.assertEqual(
            module._build_unit_project_name(
                "福泰花苑屋顶漏水维修工程",
                "福泰花苑屋顶漏水维修工程-7幢",
            ),
            "福泰花苑屋顶漏水维修工程-7幢",
        )
        self.assertEqual(
            module._build_unit_project_name(
                "平湖滨水广场消火栓主管漏水，泵房压力下降及更换消防报警主机项目",
                "单项工程-安装",
            ),
            "平湖滨水广场消火栓主管漏水，泵房压力下降及更换消防报警主机项目-单项工程-安装",
        )
        self.assertEqual(
            module._classification_cache_subject("福泰花苑屋顶漏水维修工程-7幢"),
            module._classification_cache_subject("福泰花苑屋顶漏水维修工程-8幢"),
        )

        units = module._group_classification_units(
            {
                "consultation_project_name": "福泰花苑屋顶漏水维修工程",
                "renovation_content": "屋顶漏水维修",
                "file_name": "roof.pdf",
                "consultation_time": "2026-01-01",
                "location": "浙江省嘉兴市",
            },
            [
                {
                    "page_no": 1,
                    "seq": 1,
                    "sub_project_id": "福泰花苑屋顶漏水维修工程-7幢",
                    "project_code": "0101",
                    "project_name": "屋面防水",
                    "total_price": 100,
                }
            ],
        )
        self.assertEqual(units[0]["unit_project_name"], "福泰花苑屋顶漏水维修工程-7幢")
        self.assertEqual(units[0]["ocr_values"]["unit_project_name"], "福泰花苑屋顶漏水维修工程-7幢")

    def test_batch_script_strips_only_trailing_group_project_codes(self):
        script = ROOT / "scripts" / "batch_classify_excel.py"
        spec = importlib.util.spec_from_file_location("batch_classify_excel", script)
        module = importlib.util.module_from_spec(spec)
        self.assertIsNotNone(spec)
        self.assertIsNotNone(spec.loader)
        spec.loader.exec_module(module)

        project_codes = {"10-2-37", "12-3-60", "10-1-60", "10-1-62", "10-2-39", "9-1-25"}
        self.assertEqual(
            module._strip_trailing_project_codes(
                "单项工程-安装-10-2-37-12-3-60-10-1-60-10-1-62-10-2-39-9-1-25",
                project_codes,
            ),
            "单项工程-安装",
        )
        self.assertEqual(
            module._strip_trailing_project_codes("单项工程-10-2-37-安装-12-3-60", project_codes),
            "单项工程-10-2-37-安装",
        )
        self.assertEqual(
            module._strip_trailing_project_codes("单项工程-安装-99-9-99", project_codes),
            "单项工程-安装-99-9-99",
        )

        for separator in ("-", "_", "－", "—"):
            self.assertEqual(
                module._strip_trailing_project_codes(f"单项工程-安装{separator}10-2-37", project_codes),
                "单项工程-安装",
            )

    def test_prepare_merged_items_strips_codes_from_whole_group(self):
        script = ROOT / "scripts" / "batch_classify_excel.py"
        spec = importlib.util.spec_from_file_location("batch_classify_excel", script)
        module = importlib.util.module_from_spec(spec)
        self.assertIsNotNone(spec)
        self.assertIsNotNone(spec.loader)
        spec.loader.exec_module(module)

        polluted_sub_project_id = "单项工程-安装-10-2-37-12-3-60-10-1-60-10-1-62-10-2-39-9-1-25"
        items = [
            {
                "page_no": 1,
                "seq": index,
                "sub_project_id": polluted_sub_project_id,
                "project_code": code,
                "project_name": f"清单{index}",
                "total_price": index * 100,
            }
            for index, code in enumerate(
                ["10-2-39", "10-1-62", "10-1-60", "12-3-60", "9-1-25", "10-2-37"],
                start=1,
            )
        ]

        prepared = module._prepare_merged_items(items, "龙漱湾悦居消防维修工程")

        self.assertEqual(len(prepared), 6)
        self.assertTrue(all(item["sub_project_id"] == "单项工程-安装" for item in prepared))

    def test_batch_script_cleans_ocr_project_code_pollution_from_cache_subject_and_rows(self):
        script = ROOT / "scripts" / "batch_classify_excel.py"
        spec = importlib.util.spec_from_file_location("batch_classify_excel", script)
        module = importlib.util.module_from_spec(spec)
        self.assertIsNotNone(spec)
        self.assertIsNotNone(spec.loader)
        spec.loader.exec_module(module)

        consultation_name = "龙漱湾悦居消防维修工程"
        polluted_sub_project_id = "单项工程-安装-10-2-37-12-3-60-10-1-60-10-1-62-10-2-39-9-1-25"
        unit_project_name = f"{consultation_name}-单项工程-安装"
        project_codes = ["10-2-39", "10-1-62", "10-1-60", "12-3-60", "9-1-25", "10-2-37"]
        rows = [
            {
                "page_no": 1,
                "seq": index,
                "sub_project_id": polluted_sub_project_id,
                "project_code": code,
                "project_name": f"消防维修清单{index}",
                "total_price": index * 100,
            }
            for index, code in enumerate(project_codes, start=1)
        ]
        calls: list[str] = []

        def fake_classify(project_text: str, **_kwargs) -> dict[str, object]:
            calls.append(project_text)
            return {
                "project_name": project_text,
                "project_name_text": "消防维修",
                "catalog_id": "CF-028-00",
                "category": "消防系统",
                "item": "未明确具体子项",
                "pipeline_status": "ok",
            }

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            input_file = tmp_path / "ocr_2.xlsx"
            output_file = tmp_path / "output.xlsx"
            input_file.write_bytes(
                make_ocr_workbook(
                    [
                        (
                            "ocr_2.pdf",
                            consultation_name,
                            "2026-01-01",
                            "消防维修",
                            json.dumps(rows, ensure_ascii=False),
                            "浙江省嘉兴市",
                        )
                    ]
                )
            )

            text_stdout = StringIO()
            with redirect_stdout(text_stdout):
                stats = module.classify_workbook(input_file, output_file, fake_classify, {})

            self.assertEqual(stats, (1, 0, 1, 0))
            self.assertEqual(calls, ["单项工程-安装"])

            log_text = text_stdout.getvalue()
            self.assertIn(
                f"[ROW ] ocr_2.xlsx:2 {unit_project_name} cache_subject={unit_project_name}",
                log_text,
            )
            for code in ("10-2-37", "12-3-60", "9-1-25"):
                self.assertNotIn(code, log_text)

            workbook = openpyxl.load_workbook(output_file)
            worksheet = workbook.active
            self.assertEqual(worksheet.cell(row=2, column=1).value, unit_project_name)
            self.assertEqual(worksheet.cell(row=2, column=17).value, "单项工程-安装")
            self.assertEqual(worksheet.cell(row=2, column=21).value, unit_project_name)

            output_items = json.loads(worksheet.cell(row=2, column=18).value)
            self.assertTrue(all(item["sub_project_id"] == "单项工程-安装" for item in output_items))
            for code in ("10-2-37", "12-3-60", "9-1-25"):
                self.assertNotIn(code, worksheet.cell(row=2, column=1).value)
                self.assertNotIn(code, worksheet.cell(row=2, column=17).value)
                self.assertNotIn(code, worksheet.cell(row=2, column=21).value)

    def test_batch_script_merges_ocr_pages_cleans_sub_project_and_filters_summary(self):
        script = ROOT / "scripts" / "batch_classify_excel.py"
        spec = importlib.util.spec_from_file_location("batch_classify_excel", script)
        module = importlib.util.module_from_spec(spec)
        self.assertIsNotNone(spec)
        self.assertIsNotNone(spec.loader)
        spec.loader.exec_module(module)

        calls: list[tuple[str, dict[str, object]]] = []

        def item(seq: int, name: str, code: str = "") -> dict[str, object]:
            return {
                "page_no": 1 if seq <= 4 or seq == 14 else 2,
                "seq": seq,
                "sub_project_id": "单项工程-安装-030404035001",
                "project_code": code,
                "project_name": name,
                "total_price": seq * 10,
            }

        first_page = [
            item(1, "监控摄像设备", "030404035001"),
            item(2, "录像设备"),
            item(3, "存储设备"),
            item(4, "交换机"),
            item(14, "安全防范分系统调试"),
        ]
        second_page = [
            item(5, "双绞线缆"),
            item(6, "配线"),
            item(7, "小电器"),
            item(8, "光缆"),
            item(9, "光纤连接"),
            item(10, "收发器"),
            item(11, "分线接线箱（盒）"),
            item(12, "配管"),
            item(13, "插座"),
            item(99, "脚手架搭拆"),
        ]

        def fake_classify(project_text: str, **kwargs) -> dict[str, object]:
            calls.append((project_text, kwargs))
            return {
                "project_name": project_text,
                "project_name_text": "监控系统安装",
                "catalog_id": "CF-018-02",
                "category": "弱电系统",
                "item": "视频监控",
                "pipeline_status": "ok",
            }

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            input_file = tmp_path / "input.xlsx"
            output_file = tmp_path / "output.xlsx"
            input_file.write_bytes(
                make_ocr_workbook(
                    [
                        (
                            "monitor.pdf",
                            "平湖万家花城兴和苑地下非机动车库充停场所加装监控管理系统工程",
                            "2026-01-01",
                            "加装监控管理系统",
                            json.dumps(first_page, ensure_ascii=False),
                            "浙江省嘉兴市",
                        ),
                        (
                            "monitor.pdf",
                            "平湖万家花城兴和苑地下非机动车库充停场所加装监控管理系统工程",
                            "2026-01-01",
                            "加装监控管理系统",
                            json.dumps(second_page, ensure_ascii=False),
                            "浙江省嘉兴市",
                        ),
                    ]
                )
            )

            cache: dict[str, dict[str, object]] = {}
            stats = module.classify_workbook(input_file, output_file, fake_classify, cache)

            self.assertEqual(stats, (1, 0, 1, 0))
            self.assertEqual(len(calls), 1)
            self.assertEqual(calls[0][0], "单项工程-安装")
            self.assertEqual(
                calls[0][1]["item_summary"],
                [
                    "监控摄像设备",
                    "录像设备",
                    "存储设备",
                    "交换机",
                    "安全防范分系统调试",
                    "双绞线缆",
                    "配线",
                    "小电器",
                    "光缆",
                    "光纤连接",
                    "收发器",
                    "分线接线箱（盒）",
                    "配管",
                    "插座",
                ],
            )
            self.assertNotIn("脚手架搭拆", calls[0][1]["item_summary"])
            self.assertEqual(list(cache), [next(iter(cache))])
            cache_key = next(iter(cache))
            self.assertTrue(cache_key.startswith(f"{module.CLASSIFICATION_CACHE_VERSION}:"))
            self.assertNotIn("监控摄像设备", cache_key)
            self.assertNotIn("脚手架搭拆", cache_key)

            workbook = openpyxl.load_workbook(output_file)
            worksheet = workbook.active
            self.assertEqual(worksheet.max_row, 2)
            self.assertEqual(worksheet.cell(row=2, column=17).value, "单项工程-安装")
            output_items = json.loads(worksheet.cell(row=2, column=18).value)
            self.assertEqual([row["seq"] for row in output_items], [1, 2, 3, 4, 14, 5, 6, 7, 8, 9, 10, 11, 12, 13, 99])
            self.assertTrue(all(row["sub_project_id"] == "单项工程-安装" for row in output_items))

    def test_batch_script_outputs_multiple_rows_for_multiple_sub_projects(self):
        script = ROOT / "scripts" / "batch_classify_excel.py"
        spec = importlib.util.spec_from_file_location("batch_classify_excel", script)
        module = importlib.util.module_from_spec(spec)
        self.assertIsNotNone(spec)
        self.assertIsNotNone(spec.loader)
        spec.loader.exec_module(module)

        calls: list[str] = []

        def fake_classify(project_text: str, **_kwargs) -> dict[str, object]:
            calls.append(project_text)
            return {
                "project_name": project_text,
                "project_name_text": project_text,
                "catalog_id": f"ID-{len(calls)}",
                "category": "一级",
                "item": "二级",
                "pipeline_status": "ok",
            }

        rows = [
            {
                "page_no": 1,
                "seq": 1,
                "sub_project_id": "卓越雅苑一期第一批公区维修工程-屋面渗漏维修",
                "project_code": "0101",
                "project_name": "屋面防水卷材",
                "total_price": 100,
            },
            {
                "page_no": 1,
                "seq": 2,
                "sub_project_id": "卓越雅苑一期第一批公区维修工程-外墙渗漏维修",
                "project_code": "0201",
                "project_name": "外墙涂料",
                "total_price": 200,
            },
        ]

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            input_file = tmp_path / "input.xlsx"
            output_file = tmp_path / "output.xlsx"
            input_file.write_bytes(
                make_ocr_workbook(
                    [
                        (
                            "zhuoyue.pdf",
                            "卓越雅苑一期第一批公区维修工程",
                            "2026-01-01",
                            "公区维修",
                            json.dumps(rows, ensure_ascii=False),
                            "浙江省嘉兴市",
                        )
                    ]
                )
            )

            stats = module.classify_workbook(input_file, output_file, fake_classify, {})

            self.assertEqual(stats, (2, 0, 2, 0))
            self.assertEqual(
                calls,
                [
                    "卓越雅苑一期第一批公区维修工程-屋面渗漏维修",
                    "卓越雅苑一期第一批公区维修工程-外墙渗漏维修",
                ],
            )
            workbook = openpyxl.load_workbook(output_file)
            worksheet = workbook.active
            self.assertEqual(worksheet.cell(row=2, column=17).value, calls[0])
            self.assertEqual(worksheet.cell(row=3, column=17).value, calls[1])
            self.assertEqual(worksheet.cell(row=2, column=3).value, "ID-1")
            self.assertEqual(worksheet.cell(row=3, column=3).value, "ID-2")

    def test_filter_required_ocr_rows_cli_splits_cleaned_and_removed_rows(self):
        script = ROOT / "scripts" / "filter_required_ocr_rows.py"
        headers = [
            "row_label",
            "file_name",
            "consultation_project_name",
            "consultation_time",
            "renovation_content",
            "sub_item_project_rows",
            "location",
            "note",
        ]
        rows = [
            ["ok", "file-a.pdf", "屋面工程", "2026-01-01", "防水维修", "[]", "浙江省嘉兴市", "keep"],
            ["missing-two", "file-b.pdf", "", "2026-01-02", "外墙维修", None, "浙江省嘉兴市", "remove"],
            ["missing-two-more", " \n\t ", "电梯工程", "NaN", "钢丝绳更换", "[]", "浙江省嘉兴市", "remove"],
        ]

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            input_file = tmp_path / "ocr.xlsx"
            clean_output = tmp_path / "cleaned.xlsx"
            removed_output = tmp_path / "removed.xlsx"
            input_file.write_bytes(make_workbook_with_headers(headers, rows))

            run = subprocess.run(
                [
                    sys.executable,
                    str(script),
                    str(input_file),
                    "--clean-output",
                    str(clean_output),
                    "--removed-output",
                    str(removed_output),
                    "--overwrite",
                ],
                cwd=ROOT,
                check=False,
                capture_output=True,
                text=True,
                timeout=10,
            )

            self.assertEqual(run.returncode, 0, run.stdout + run.stderr)
            self.assertIn("[DONE] input rows: 3", run.stdout)
            self.assertIn("[DONE] cleaned rows: 1", run.stdout)
            self.assertIn("[DONE] removed rows: 2", run.stdout)
            self.assertIn("  file_name: 1", run.stdout)
            self.assertIn("  consultation_project_name: 1", run.stdout)
            self.assertIn("  consultation_time: 1", run.stdout)
            self.assertIn("  renovation_content: 0", run.stdout)
            self.assertIn("  sub_item_project_rows: 1", run.stdout)
            self.assertIn("  location: 0", run.stdout)

            clean_workbook = openpyxl.load_workbook(clean_output, data_only=False)
            clean_sheet = clean_workbook.active
            clean_headers = [clean_sheet.cell(row=1, column=i).value for i in range(1, clean_sheet.max_column + 1)]
            self.assertEqual(clean_headers, headers)
            self.assertEqual(clean_sheet.max_row, 2)
            self.assertEqual([clean_sheet.cell(row=2, column=i).value for i in range(1, 9)], rows[0])

            removed_workbook = openpyxl.load_workbook(removed_output, data_only=False)
            removed_sheet = removed_workbook.active
            removed_headers = [
                removed_sheet.cell(row=1, column=i).value for i in range(1, removed_sheet.max_column + 1)
            ]
            self.assertEqual(
                removed_headers,
                ["source_row_id", "missing_required_fields", "removed_reason"] + headers,
            )
            self.assertEqual(removed_sheet.max_row, 3)
            self.assertEqual(removed_sheet.cell(row=2, column=1).value, 3)
            self.assertEqual(
                removed_sheet.cell(row=2, column=2).value,
                "consultation_project_name,sub_item_project_rows",
            )
            self.assertEqual(removed_sheet.cell(row=2, column=3).value, "缺少必填 OCR 字段")
            self.assertEqual(removed_sheet.cell(row=2, column=4).value, "missing-two")
            self.assertEqual(removed_sheet.cell(row=3, column=1).value, 4)
            self.assertEqual(removed_sheet.cell(row=3, column=2).value, "file_name,consultation_time")
            self.assertEqual(removed_sheet.cell(row=3, column=4).value, "missing-two-more")

    def test_filter_required_ocr_rows_empty_value_detection(self):
        script = ROOT / "scripts" / "filter_required_ocr_rows.py"
        spec = importlib.util.spec_from_file_location("filter_required_ocr_rows", script)
        module = importlib.util.module_from_spec(spec)
        self.assertIsNotNone(spec)
        self.assertIsNotNone(spec.loader)
        spec.loader.exec_module(module)

        empty_values = [None, "", "   ", "\n\t ", "null", "None", "nan", "NaN"]
        for value in empty_values:
            with self.subTest(value=value):
                self.assertTrue(module.is_empty_required_value(value))
        self.assertFalse(module.is_empty_required_value("0"))
        self.assertFalse(module.is_empty_required_value(0))

    def test_filter_required_ocr_rows_missing_headers_fail_without_outputs(self):
        script = ROOT / "scripts" / "filter_required_ocr_rows.py"
        headers = [
            "file_name",
            "consultation_project_name",
            "consultation_time",
            "renovation_content",
            "sub_item_project_rows",
        ]
        rows = [["file-a.pdf", "屋面工程", "2026-01-01", "防水维修", "[]"]]

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            input_file = tmp_path / "ocr.xlsx"
            clean_output = tmp_path / "cleaned.xlsx"
            removed_output = tmp_path / "removed.xlsx"
            input_file.write_bytes(make_workbook_with_headers(headers, rows))

            run = subprocess.run(
                [
                    sys.executable,
                    str(script),
                    str(input_file),
                    "--clean-output",
                    str(clean_output),
                    "--removed-output",
                    str(removed_output),
                    "--overwrite",
                ],
                cwd=ROOT,
                check=False,
                capture_output=True,
                text=True,
                timeout=10,
            )

            self.assertEqual(run.returncode, 1)
            self.assertIn("[ERROR] 输入 Excel 缺少必要列: location", run.stdout)
            self.assertFalse(clean_output.exists())
            self.assertFalse(removed_output.exists())

    def test_filter_required_ocr_rows_requires_overwrite_for_existing_outputs(self):
        script = ROOT / "scripts" / "filter_required_ocr_rows.py"

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            input_file = tmp_path / "ocr.xlsx"
            clean_output = tmp_path / "cleaned.xlsx"
            removed_output = tmp_path / "removed.xlsx"
            input_file.write_bytes(
                make_ocr_workbook([("file-a.pdf", "屋面工程", "2026-01-01", "防水维修", "[]", "浙江省嘉兴市")])
            )
            clean_output.write_text("existing", encoding="utf-8")

            run = subprocess.run(
                [
                    sys.executable,
                    str(script),
                    str(input_file),
                    "--clean-output",
                    str(clean_output),
                    "--removed-output",
                    str(removed_output),
                ],
                cwd=ROOT,
                check=False,
                capture_output=True,
                text=True,
                timeout=10,
            )

            self.assertEqual(run.returncode, 1)
            self.assertIn("[ERROR] 输出已存在，请加 --overwrite 或更换输出路径", run.stdout)
            self.assertEqual(clean_output.read_text(encoding="utf-8"), "existing")
            self.assertFalse(removed_output.exists())

    def test_batch_script_importable(self):
        script = ROOT / "scripts" / "batch_classify_excel.py"
        spec = importlib.util.spec_from_file_location("batch_classify_excel", script)
        module = importlib.util.module_from_spec(spec)
        self.assertIsNotNone(spec)
        self.assertIsNotNone(spec.loader)
        spec.loader.exec_module(module)
        self.assertEqual(module.RESULT_HEADERS, NEW_HEADERS)


if __name__ == "__main__":
    unittest.main()
