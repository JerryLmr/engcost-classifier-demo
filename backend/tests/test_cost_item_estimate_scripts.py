from __future__ import annotations

import importlib.util
import sys
import tempfile
import types
import unittest
from pathlib import Path
from unittest.mock import patch

import openpyxl

try:
    import numpy as np
    import pandas as pd
except ImportError:
    np = None
    pd = None


ROOT = Path(__file__).resolve().parents[2]


def load_script_module(name: str, relative_path: str):
    spec = importlib.util.spec_from_file_location(name, ROOT / relative_path)
    module = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    sys.modules[name] = module
    spec.loader.exec_module(module)
    return module


build_samples_script = load_script_module("build_cost_item_samples", "scripts/build_cost_item_samples.py")
run_ingest_batch = load_script_module("run_ingest_batch", "scripts/run_ingest_batch.py")
merge_samples = load_script_module("merge_cost_item_sample_batches", "scripts/merge_cost_item_sample_batches.py")

if np is not None and pd is not None:
    build_index = load_script_module("build_cost_item_embedding_index", "scripts/build_cost_item_embedding_index.py")
    query_estimate_llm = load_script_module("query_cost_estimate_llm", "scripts/query_cost_estimate_llm.py")
else:
    build_index = None
    query_estimate_llm = None


@unittest.skipIf(np is None or pd is None, "cost item estimate dependencies are not installed")
class CostItemEstimateScriptTestCase(unittest.TestCase):
    def raw_sample_frame(self) -> pd.DataFrame:
        return pd.DataFrame(
            [
                {
                    "project_key": "batch-a::2",
                    "batch_id": "batch-a",
                    "source_row_id": 2,
                    "item_row_id": "2-1",
                    "工程名称": "屋面漏水维修工程",
                    "project_name_text": "屋面漏水维修",
                    "catalog_id": "CP-002-03",
                    "一级分类": "屋面",
                    "二级分类": "防水层",
                    "维修状态": "维修",
                    "标准对象": "共用部位",
                    "是否复合工程": "否",
                    "复合目录": "",
                    "seq": 1,
                    "cost_item_name": "屋面卷材防水",
                    "project_description": "3.0mm SBS 沥青防水卷材",
                    "unit": "平方米",
                    "unit_normalized": "m²",
                    "quantity": 100,
                    "unit_price": 80,
                    "total_price": 8000,
                    "labor_unit_price": 20,
                    "machinery_unit_price": 5,
                    "consultation_time": "2026-03-01",
                    "location": "浙江省嘉兴市",
                    "stable_sample_id": "stable-2-1",
                },
                {
                    "project_key": "batch-a::2",
                    "batch_id": "batch-a",
                    "source_row_id": 2,
                    "item_row_id": "2-2",
                    "工程名称": "屋面漏水维修工程",
                    "project_name_text": "屋面漏水维修",
                    "catalog_id": "CP-002-03",
                    "一级分类": "屋面",
                    "二级分类": "防水层",
                    "维修状态": "维修",
                    "标准对象": "共用部位",
                    "seq": 2,
                    "cost_item_name": "防水层拆除",
                    "project_description": "拆除原屋面防水层",
                    "unit": "平方米",
                    "unit_normalized": "m²",
                    "quantity": 100,
                    "unit_price": 20,
                    "total_price": 2000,
                    "labor_unit_price": 12,
                    "machinery_unit_price": 1,
                    "consultation_time": "2026-03-01",
                    "location": "浙江省嘉兴市",
                },
                {
                    "project_key": "batch-a::3",
                    "batch_id": "batch-a",
                    "source_row_id": 3,
                    "item_row_id": "3-1",
                    "工程名称": "外墙维修工程",
                    "project_name_text": "外墙维修",
                    "catalog_id": "CP-003-01",
                    "一级分类": "外墙面",
                    "二级分类": "面层",
                    "维修状态": "维修",
                    "标准对象": "共用部位",
                    "seq": 1,
                    "cost_item_name": "外墙防水",
                    "project_description": "外墙渗漏处理",
                    "unit": "平方米",
                    "unit_normalized": "m²",
                    "quantity": 200,
                    "unit_price": 100,
                    "total_price": 20000,
                    "labor_unit_price": 30,
                    "machinery_unit_price": 8,
                    "consultation_time": "2024-01-01",
                    "location": "浙江省杭州市",
                },
                {
                    "project_key": "batch-a::4",
                    "batch_id": "batch-a",
                    "source_row_id": 4,
                    "item_row_id": "4-1",
                    "工程名称": "管道维修工程",
                    "project_name_text": "管道维修",
                    "catalog_id": "CF-015-04",
                    "一级分类": "给排水系统",
                    "二级分类": "管道",
                    "维修状态": "维修",
                    "标准对象": "共用设施设备",
                    "seq": 1,
                    "cost_item_name": "管道更换",
                    "project_description": "DN100 镀锌钢管更换",
                    "unit": "米",
                    "unit_normalized": "m",
                    "quantity": 30,
                    "unit_price": 100,
                    "total_price": 3000,
                    "labor_unit_price": "",
                    "machinery_unit_price": "",
                    "consultation_time": "2026-02-01",
                    "location": "浙江省嘉兴市",
                },
            ]
        )

    def prepared_samples(self) -> pd.DataFrame:
        samples = build_index.normalize_numeric_columns(
            build_index.ensure_optional_columns(build_index.ensure_project_key(self.raw_sample_frame()))
        )
        samples.insert(0, "sample_index", range(len(samples)))
        samples["project_package_id"] = samples["project_key"]
        samples["item_retrieval_text"] = samples.apply(build_index.build_item_retrieval_text, axis=1)
        samples["fine_signature"] = samples.apply(build_index.build_fine_signature, axis=1)
        samples["family_signature"] = samples.apply(build_index.build_family_signature, axis=1)
        return samples

    def write_samples_xlsx(self, path: Path, frame: pd.DataFrame) -> None:
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            frame.to_excel(writer, sheet_name="samples", index=False)

    def test_normalize_embeddings_handles_zero_vector(self):
        embeddings = np.array([[3.0, 4.0], [0.0, 0.0]], dtype=np.float32)
        normalized = build_index.normalize_embeddings(embeddings)
        self.assertEqual(normalized.dtype, np.float32)
        self.assertAlmostEqual(float(np.linalg.norm(normalized[0])), 1.0)
        self.assertEqual(normalized[1].tolist(), [0.0, 0.0])

    def test_shared_normalize_unit_handles_square_and_cubic_units(self):
        self.assertEqual(build_samples_script.normalize_unit("m^2"), "m²")
        self.assertEqual(build_samples_script.normalize_unit("平方米"), "m²")
        self.assertEqual(build_samples_script.normalize_unit("平"), "m²")
        self.assertEqual(build_samples_script.normalize_unit("m^{3}"), "m³")
        self.assertEqual(build_samples_script.normalize_unit(" 台 "), "台")

    def test_load_samples_generates_package_fields_and_project_key_fallback(self):
        samples = self.raw_sample_frame().drop(columns=["project_key", "stable_sample_id"])
        samples.loc[0, "source_row_id"] = 2.0
        samples = samples.drop(columns=["labor_unit_price"])

        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "samples.xlsx"
            self.write_samples_xlsx(path, samples)
            loaded = build_index.load_samples(path)

        self.assertEqual(loaded["project_key"].iloc[0], "batch-a::2")
        self.assertEqual(loaded["sample_index"].tolist(), [0, 1, 2, 3])
        self.assertEqual(loaded["project_package_id"].iloc[0], "batch-a::2")
        self.assertTrue(loaded["labor_unit_price"].isna().all())
        self.assertIn("清单项：屋面卷材防水", loaded["item_retrieval_text"].iloc[0])
        self.assertIn("项目特征：3.0mm SBS 沥青防水卷材", loaded["item_retrieval_text"].iloc[0])
        self.assertEqual(loaded["fine_signature"].iloc[0], "屋面卷材防水 | 3.0mm sbs 沥青防水卷材 | m²")
        self.assertEqual(loaded["family_signature"].iloc[0], "屋面卷材防水 | m²")

    def test_load_samples_rejects_missing_required_columns(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "samples.xlsx"
            self.write_samples_xlsx(path, pd.DataFrame([{"project_key": "a"}]))

            with self.assertRaisesRegex(ValueError, "samples sheet 缺少必要字段"):
                build_index.load_samples(path)

    def test_build_project_packages_aggregates_project_key_without_section_split(self):
        packages = build_index.build_project_packages(self.prepared_samples())

        self.assertEqual(packages["project_package_id"].tolist(), ["batch-a::2", "batch-a::3", "batch-a::4"])
        self.assertEqual(packages.columns.tolist(), build_index.PROJECT_PACKAGE_COLUMNS)
        roof = packages[packages["project_package_id"] == "batch-a::2"].iloc[0]
        self.assertEqual(roof["project_package_title"], "屋面漏水维修工程 / 屋面漏水维修")
        self.assertEqual(roof["item_count"], 2)
        self.assertIn("CP-002-03 | 屋面 / 防水层 / 维修 / 共用部位", roof["catalog_summary"])
        self.assertIn("屋面卷材防水：3.0mm SBS 沥青防水卷材；单位：平方米", roof["item_summary"])
        self.assertIn("包含清单：", roof["package_text"])
        self.assertIn("- 防水层拆除：拆除原屋面防水层；单位：平方米", roof["package_text"])

    def test_build_index_meta_tracks_new_files_only(self):
        meta = build_index.build_index_meta(Path("samples.xlsx"), "demo-model", 4, 3, 2)

        self.assertEqual(
            meta["files"],
            {
                "samples": "samples.parquet",
                "project_packages": "project_packages.parquet",
                "project_package_embeddings": "project_package_embeddings.npy",
                "item_embeddings": "item_embeddings.npy",
            },
        )
        self.assertEqual(meta["package_count"], 3)
        self.assertNotIn("project_groups", meta["files"])
        self.assertIn("fine_signature", meta["field_descriptions"])

    def test_write_index_outputs_new_files_and_removes_legacy_files(self):
        samples = self.prepared_samples()
        packages = build_index.build_project_packages(samples)
        package_embeddings = np.ones((len(packages), 2), dtype=np.float32)
        item_embeddings = np.ones((len(samples), 2), dtype=np.float32)
        meta = build_index.build_index_meta(Path("samples.xlsx"), "demo-model", len(samples), len(packages), 2)

        with tempfile.TemporaryDirectory() as tmpdir:
            output_dir = Path(tmpdir) / "index"
            output_dir.mkdir()
            for name in build_index.LEGACY_OUTPUT_FILES:
                (output_dir / name).write_text("legacy", encoding="utf-8")

            build_index.write_index(samples, packages, package_embeddings, item_embeddings, output_dir, meta)

            self.assertTrue((output_dir / "samples.parquet").exists())
            self.assertTrue((output_dir / "project_packages.parquet").exists())
            self.assertTrue((output_dir / "project_package_embeddings.npy").exists())
            self.assertTrue((output_dir / "item_embeddings.npy").exists())
            self.assertTrue((output_dir / "index_meta.json").exists())
            for name in build_index.LEGACY_OUTPUT_FILES:
                self.assertFalse((output_dir / name).exists())

    def test_build_index_validate_output_dir_requires_overwrite_for_existing_directory(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            output_dir = Path(tmpdir) / "index"
            output_dir.mkdir()

            with self.assertRaisesRegex(ValueError, "输出已存在，请加 --overwrite"):
                build_index.validate_output_dir(output_dir, overwrite=False)
            build_index.validate_output_dir(output_dir, overwrite=True)

    def test_build_index_parse_args_defaults_to_merged_samples(self):
        with patch.object(sys, "argv", ["build_cost_item_embedding_index.py"]):
            args = build_index.parse_args()

        self.assertEqual(args.samples, "samples/cost_item_samples_all.xlsx")
        self.assertEqual(args.output_dir, "embeddings")
        self.assertEqual(args.model, "BAAI/bge-m3")
        self.assertEqual(args.batch_size, 32)
        self.assertFalse(args.overwrite)

    def test_query_parse_args_uses_new_top_package_and_item_options(self):
        with patch.object(sys, "argv", ["query_cost_estimate_llm.py", "--text", "屋面漏水"]):
            args = query_estimate_llm.parse_args()

        self.assertEqual(args.index_dir, "embeddings")
        self.assertEqual(args.top_packages, 20)
        self.assertEqual(args.top_items, 300)
        self.assertFalse(hasattr(args, "top_k"))
        self.assertFalse(hasattr(args, "project_name_weight"))
        self.assertFalse(hasattr(args, "project_detail_weight"))

    def test_query_load_index_reads_new_files_and_validates_shapes(self):
        samples = self.prepared_samples()
        packages = build_index.build_project_packages(samples)
        with tempfile.TemporaryDirectory() as tmpdir:
            index_dir = Path(tmpdir)
            samples.to_parquet(index_dir / "samples.parquet", index=False)
            packages.to_parquet(index_dir / "project_packages.parquet", index=False)
            np.save(index_dir / "project_package_embeddings.npy", np.ones((len(packages), 2), dtype=np.float32))
            np.save(index_dir / "item_embeddings.npy", np.zeros((len(samples), 2), dtype=np.float32))
            (index_dir / "index_meta.json").write_text('{"model": "demo-model"}', encoding="utf-8")

            loaded_samples, loaded_packages, package_embeddings, item_embeddings, meta = query_estimate_llm.load_index(index_dir)

        self.assertEqual(len(loaded_samples), 4)
        self.assertEqual(len(loaded_packages), 3)
        self.assertEqual(package_embeddings.shape, (3, 2))
        self.assertEqual(item_embeddings.shape, (4, 2))
        self.assertEqual(meta["model"], "demo-model")

    def test_query_load_index_rejects_legacy_only_index(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            index_dir = Path(tmpdir)
            self.prepared_samples().to_parquet(index_dir / "samples.parquet", index=False)
            (index_dir / "project_groups.parquet").write_text("legacy", encoding="utf-8")
            np.save(index_dir / "project_name_embeddings.npy", np.ones((3, 2), dtype=np.float32))
            np.save(index_dir / "project_detail_embeddings.npy", np.ones((3, 2), dtype=np.float32))
            np.save(index_dir / "item_text_embeddings.npy", np.ones((4, 2), dtype=np.float32))
            (index_dir / "index_meta.json").write_text("{}", encoding="utf-8")

            with self.assertRaisesRegex(ValueError, "索引目录缺少文件"):
                query_estimate_llm.load_index(index_dir)

    def test_llm_query_load_embedding_model_forces_cpu(self):
        fake_module = types.ModuleType("sentence_transformers")
        calls = []

        class FakeSentenceTransformer:
            def __init__(self, model_name, device=None):
                calls.append((model_name, device))

        fake_module.SentenceTransformer = FakeSentenceTransformer

        with patch.dict(sys.modules, {"sentence_transformers": fake_module}):
            model = query_estimate_llm.load_embedding_model("demo-model")

        self.assertIsInstance(model, FakeSentenceTransformer)
        self.assertEqual(calls, [("demo-model", "cpu")])

    def test_understand_query_uses_llm_and_falls_back_on_failure(self):
        llm_result = {
            "semantic_query_text": "屋面漏水 3mm SBS 防水",
            "need_summary": "屋面漏水防水维修",
            "known_constraints": {"部位": "屋面", "用户明确数量": "500平"},
            "likely_catalog": {"一级分类": "屋面", "二级分类": "防水层"},
            "calculation_notes": "面积约500平",
        }
        with patch.object(query_estimate_llm, "request_llm_json", return_value=llm_result):
            parsed = query_estimate_llm.understand_query("屋面漏水，500平")

        self.assertTrue(parsed.llm_success)
        self.assertEqual(parsed.semantic_query_text, "屋面漏水 3mm SBS 防水")
        self.assertEqual(parsed.known_constraints["部位"], "屋面")
        self.assertEqual(parsed.likely_catalog["二级分类"], "防水层")

        with patch.object(query_estimate_llm, "request_llm_json", side_effect=query_estimate_llm.LLMServiceError("down")):
            fallback = query_estimate_llm.understand_query("屋面漏水")

        self.assertFalse(fallback.llm_success)
        self.assertEqual(fallback.semantic_query_text, "屋面漏水")
        self.assertIn("LLM 查询理解失败", fallback.parse_notes[0])

    def test_package_and_item_recall_return_ranked_rows(self):
        samples = self.prepared_samples()
        packages = build_index.build_project_packages(samples)
        package_embeddings = np.array([[1.0, 0.0], [0.2, 0.8], [0.0, 1.0]], dtype=np.float32)
        item_scores = np.array([0.9, 0.3, 0.5, 0.1], dtype=np.float32)

        matched = query_estimate_llm.score_project_packages(
            packages,
            package_embeddings,
            np.array([1.0, 0.0], dtype=np.float32),
            top_packages=2,
        )
        direct = query_estimate_llm.score_direct_items(samples, item_scores, top_items=2)

        self.assertEqual(matched["project_package_id"].tolist(), ["batch-a::2", "batch-a::3"])
        self.assertEqual(matched["rank"].tolist(), [1, 2])
        self.assertEqual(direct["sample_index"].tolist(), [0, 2])
        self.assertEqual(direct["item_score"].tolist(), [0.8999999761581421, 0.5])

    def test_candidate_pool_merges_package_and_direct_hits_with_scores(self):
        samples = self.prepared_samples()
        packages = build_index.build_project_packages(samples)
        matched = packages[packages["project_package_id"].isin(["batch-a::2", "batch-a::3"])].copy()
        matched.insert(0, "package_score", [0.9, 0.6])
        matched.insert(0, "rank", [1, 2])
        direct = samples[samples["sample_index"].isin([0, 3])].copy()
        item_scores = np.array([0.95, 0.2, 0.3, 0.7], dtype=np.float32)
        understanding = query_estimate_llm.QueryUnderstanding(
            raw_query="屋面防水500平",
            semantic_query_text="屋面防水",
            need_summary="",
            known_constraints={"用户明确数量": "500 m²"},
            likely_catalog={"一级分类": "屋面", "二级分类": "防水层"},
            calculation_notes="",
            parse_notes=[],
            llm_success=True,
        )

        candidates = query_estimate_llm.candidate_pool(samples, matched, direct, item_scores, understanding)

        self.assertEqual(sorted(candidates["sample_index"].tolist()), [0, 1, 2, 3])
        roof = candidates[candidates["sample_index"] == 0].iloc[0]
        pipe = candidates[candidates["sample_index"] == 3].iloc[0]
        self.assertAlmostEqual(float(roof["package_score"]), 0.9)
        self.assertAlmostEqual(float(roof["item_score"]), 0.95)
        self.assertAlmostEqual(float(roof["cooccur_score"]), 0.5)
        self.assertEqual(float(roof["catalog_score"]), 1.0)
        self.assertEqual(float(roof["unit_score"]), 1.0)
        self.assertEqual(pipe["package_score"], 0.0)
        self.assertEqual(pipe["direct_hit"], True)
        self.assertEqual(pipe["catalog_score"], 0.2)

    def test_candidate_item_stats_keeps_different_project_descriptions_separate(self):
        samples = self.prepared_samples()
        duplicate = samples.iloc[0].copy()
        duplicate["sample_index"] = 4
        duplicate["project_package_id"] = "batch-a::5"
        duplicate["project_key"] = "batch-a::5"
        duplicate["project_description"] = "4.0mm SBS 沥青防水卷材"
        duplicate["quantity"] = 120
        duplicate["unit_price"] = 120
        duplicate["total_price"] = 14400
        duplicate["fine_signature"] = build_index.build_fine_signature(duplicate)
        samples = pd.concat([samples, pd.DataFrame([duplicate])], ignore_index=True)
        candidates = samples.copy()
        candidates["package_score"] = 0.8
        candidates["package_rank"] = 1
        candidates["item_score"] = [0.9, 0.2, 0.4, 0.1, 0.85]
        candidates["cooccur_score"] = 0.5
        candidates["catalog_score"] = 1.0
        candidates["unit_score"] = 0.5
        candidates["final_score"] = 0.7
        candidates["evidence_ref"] = candidates["sample_index"].map(lambda value: f"E{int(value)}")

        stats = query_estimate_llm.build_candidate_item_stats(candidates)

        roof_stats = stats[stats["cost_item_name"] == "屋面卷材防水"]
        self.assertEqual(len(roof_stats), 2)
        self.assertEqual(set(roof_stats["project_description"]), {"3.0mm SBS 沥青防水卷材", "4.0mm SBS 沥青防水卷材"})
        first = roof_stats[roof_stats["project_description"] == "3.0mm SBS 沥青防水卷材"].iloc[0]
        self.assertEqual(first["历史样本数"], 1)
        self.assertEqual(first["历史综合单价中位数"], 80.0)
        self.assertEqual(first["evidence_refs"], "E0")

    def test_evidence_items_preserve_project_description_and_prices(self):
        candidates = self.prepared_samples().head(1).copy()
        candidates["package_rank"] = 1
        candidates["package_score"] = 0.9
        candidates["item_score"] = 0.95
        candidates["cooccur_score"] = 1.0
        candidates["catalog_score"] = 1.0
        candidates["unit_score"] = 0.5
        candidates["final_score"] = 0.92
        candidates["evidence_ref"] = "E0"

        evidence = query_estimate_llm.build_evidence_items(candidates)

        self.assertEqual(evidence.columns.tolist(), query_estimate_llm.EVIDENCE_ITEM_COLUMNS)
        self.assertEqual(evidence.loc[0, "project_description"], "3.0mm SBS 沥青防水卷材")
        self.assertEqual(evidence.loc[0, "unit_price"], 80)
        self.assertEqual(evidence.loc[0, "来源工程名称"], "屋面漏水维修工程")

    def test_suggested_bill_success_marks_adopted_candidates(self):
        stats = pd.DataFrame(
            [
                {
                    "fine_signature": "a",
                    "family_signature": "a",
                    "cost_item_name": "屋面卷材防水",
                    "project_description": "3mm SBS",
                    "unit": "m²",
                    "历史样本数": 1,
                    "来源工程包数": 1,
                    "final_score": 0.9,
                    "是否被LLM采用": "否",
                    "evidence_refs": "E0, E1",
                }
            ]
        )
        result = {
            "suggested_bill": [
                {
                    "seq": 1,
                    "recommend_type": "直接匹配项",
                    "cost_item_name": "屋面卷材防水",
                    "project_description": "3mm SBS",
                    "unit": "m²",
                    "suggested_quantity": 500,
                    "quantity_basis": "用户明确面积",
                    "unit_price_low": 80,
                    "unit_price_mid": 100,
                    "unit_price_high": 120,
                    "estimated_amount_low": 40000,
                    "estimated_amount_mid": 50000,
                    "estimated_amount_high": 60000,
                    "adopt_reason": "直接命中",
                    "uncertainty_note": "",
                    "evidence_refs": ["E0"],
                }
            ]
        }

        bill = query_estimate_llm.suggested_bill_from_llm_result(result)
        marked = query_estimate_llm.mark_adopted_candidates(stats, bill)

        self.assertEqual(bill.columns.tolist(), query_estimate_llm.SUGGESTED_BILL_COLUMNS)
        self.assertEqual(bill.loc[0, "来源证据"], "E0")
        self.assertEqual(marked.loc[0, "是否被LLM采用"], "是")

    def test_generate_suggested_bill_fallback_records_failure(self):
        understanding = query_estimate_llm.QueryUnderstanding("屋面", "屋面", "", {}, {}, "", [], True)
        stats = pd.DataFrame(
            [
                {
                    "cost_item_name": "屋面卷材防水",
                    "project_description": "3mm SBS",
                    "unit": "m²",
                    "历史综合单价最小值": 80,
                    "历史综合单价中位数": 100,
                    "历史综合单价最大值": 120,
                    "历史合价最小值": 8000,
                    "历史合价中位数": 10000,
                    "历史合价最大值": 12000,
                    "evidence_refs": "E0",
                }
            ]
        )

        with patch.object(query_estimate_llm, "request_llm_json", side_effect=query_estimate_llm.LLMServiceError("down")):
            bill, success, fallback, error, prompt = query_estimate_llm.generate_suggested_bill(
                understanding,
                pd.DataFrame(),
                stats,
                pd.DataFrame(),
            )

        self.assertFalse(success)
        self.assertTrue(fallback)
        self.assertIn("down", error)
        self.assertIn("candidate_item_stats", prompt)
        self.assertEqual(bill.loc[0, "推荐类型"], "fallback_candidate")
        self.assertEqual(bill.loc[0, "不确定性说明"], "需人工确认；LLM suggested_bill 生成失败")

    def test_write_query_result_workbook_has_new_five_sheets(self):
        understanding = query_estimate_llm.QueryUnderstanding("屋面", "屋面", "", {}, {}, "", [], True)
        result = query_estimate_llm.QueryResult(
            understanding=understanding,
            suggested_bill=pd.DataFrame(
                [
                    {
                        "序号": 1,
                        "推荐类型": "直接匹配项",
                        "清单项名称": "屋面卷材防水",
                        "项目特征/施工工艺": "3mm SBS",
                        "单位": "m²",
                        "建议工程量": 500,
                        "工程量依据": "用户明确",
                        "综合单价低值": 80,
                        "综合单价中值": 100,
                        "综合单价高值": 120,
                        "估算金额低值": 40000,
                        "估算金额中值": 50000,
                        "估算金额高值": 60000,
                        "采用理由": "直接命中",
                        "不确定性说明": "",
                        "来源证据": "E0",
                    }
                ],
                columns=query_estimate_llm.SUGGESTED_BILL_COLUMNS,
            ),
            matched_project_packages=pd.DataFrame(columns=query_estimate_llm.MATCHED_PROJECT_PACKAGE_COLUMNS),
            candidate_item_stats=pd.DataFrame(columns=query_estimate_llm.CANDIDATE_ITEM_STATS_COLUMNS),
            evidence_items=pd.DataFrame(columns=query_estimate_llm.EVIDENCE_ITEM_COLUMNS),
            parse_info=pd.DataFrame([{"字段": "是否 fallback", "值": "否"}]),
        )

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "query_result.xlsx"
            query_estimate_llm.write_query_result_workbook(output_path, result)
            workbook = openpyxl.load_workbook(output_path, data_only=True)
            self.assertEqual(
                workbook.sheetnames,
                ["suggested_bill", "matched_project_packages", "candidate_item_stats", "evidence_items", "parse_info"],
            )
            suggested_headers = [
                workbook["suggested_bill"].cell(row=1, column=column).value
                for column in range(1, workbook["suggested_bill"].max_column + 1)
            ]
            workbook.close()

        self.assertEqual(suggested_headers, query_estimate_llm.SUGGESTED_BILL_COLUMNS)

    def test_query_validate_output_path_requires_overwrite_for_existing_output(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "query.xlsx"
            output_path.write_text("existing", encoding="utf-8")

            with self.assertRaisesRegex(ValueError, "输出已存在，请加 --overwrite"):
                query_estimate_llm.validate_output_path(output_path, overwrite=False)
            query_estimate_llm.validate_output_path(output_path, overwrite=True)

    def test_run_ingest_batch_parses_batch_id_from_filename(self):
        input_path = Path("excel_inputs/audit_ocr_export_20260630_001.xlsx")

        self.assertEqual(run_ingest_batch.infer_batch_id(input_path), "20260630_001")
        self.assertEqual(run_ingest_batch.batch_id_from_args(input_path, "manual_001"), "manual_001")

        with self.assertRaisesRegex(ValueError, "无法从文件名解析 batch_id"):
            run_ingest_batch.infer_batch_id(Path("excel_inputs/audit_ocr_export.xlsx"))

    def test_merge_samples_validate_output_paths_requires_overwrite_for_output_or_report(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            tmp_path = Path(tmpdir)
            output_path = tmp_path / "samples" / "cost_item_samples_all.xlsx"
            report_path = merge_samples.dedup_report_path(output_path)
            output_path.parent.mkdir(parents=True, exist_ok=True)
            report_path.write_text("existing", encoding="utf-8")

            with self.assertRaisesRegex(ValueError, "输出已存在，请加 --overwrite"):
                merge_samples.validate_output_paths(output_path, report_path, overwrite=False)
            merge_samples.validate_output_paths(output_path, report_path, overwrite=True)

    def test_legacy_query_entrypoint_removed_and_readme_points_to_official_script(self):
        legacy_name = "query_cost_" "item_estimate.py"
        self.assertFalse((ROOT / "scripts" / legacy_name).exists())
        readme = (ROOT / "README.md").read_text(encoding="utf-8")
        self.assertIn("scripts/query_cost_estimate_llm.py", readme)
        self.assertNotIn(f"scripts/{legacy_name}", readme)
        self.assertNotIn("project_name_embeddings.npy", readme)
        self.assertNotIn("project_detail_embeddings.npy", readme)
        self.assertNotIn("item_text_embeddings.npy", readme)


if __name__ == "__main__":
    unittest.main()
