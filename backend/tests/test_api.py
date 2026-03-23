import unittest
import sys
from io import BytesIO
from pathlib import Path
from unittest.mock import patch

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

try:
    import openpyxl
except ImportError:  # pragma: no cover
    openpyxl = None

try:
    from fastapi.testclient import TestClient
except (ImportError, RuntimeError):  # pragma: no cover
    TestClient = None

if TestClient is not None:
    from app import app


@unittest.skipIf(openpyxl is None or TestClient is None, "测试依赖未安装")
class ApiTestCase(unittest.TestCase):
    def setUp(self):
        self.client = TestClient(app)

    def test_health(self):
        response = self.client.get("/api/health")
        self.assertEqual(response.status_code, 200)
        self.assertIn("model", response.json())

    def test_single_classify(self):
        response = self.client.post("/api/classify", json={"text": "消防喷淋管网维修"})
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(data["level1"], "消防")
        self.assertEqual(data["level2"], "消防管网维修")
        self.assertIn("is_composite", data)
        self.assertIn("needs_review", data)
        self.assertIn("composite_reason", data)
        self.assertIn("secondary_candidates", data)

    def test_boundary_driven_single_classify(self):
        response = self.client.post("/api/classify", json={"text": "外墙渗漏水维修"})
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(data["level1"], "防水工程")
        self.assertEqual(data["level2"], "外墙防水")

    @patch("services.llm_client.request_llm_classification", side_effect=RuntimeError("offline"))
    def test_excel_classify(self, _mock_request):
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.cell(row=1, column=1, value="工程名称")
        worksheet.cell(row=2, column=1, value="消防喷淋管网维修")
        worksheet.cell(row=3, column=1, value="某小区公共区域综合整治提升项目")

        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        response = self.client.post(
            "/api/classify-excel",
            files={
                "file": (
                    "sample.xlsx",
                    output.getvalue(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response.headers["content-type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        workbook = openpyxl.load_workbook(BytesIO(response.content))
        worksheet = workbook.active
        headers = [worksheet.cell(row=1, column=i).value for i in range(1, worksheet.max_column + 1)]
        self.assertIn("是否复合工程", headers)
        self.assertIn("是否建议复核", headers)
        self.assertIn("复合原因", headers)
        self.assertIn("候选分类", headers)

    def test_analyze_excel(self):
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        headers = [
            "工程名称",
            "一级分类",
            "二级分类",
            "分类方式",
            "分类依据",
            "是否复合工程",
            "是否建议复核",
            "结构类型",
            "复合原因",
            "候选分类",
        ]
        for col, value in enumerate(headers, start=1):
            worksheet.cell(row=1, column=col, value=value)
        worksheet.append(["消防喷淋管网维修", "消防", "消防管网维修", "规则优先", "关键词命中", "否", "否", "single_project", "", ""])
        worksheet.append(
            [
                "某综合项目",
                "公共设施",
                "公共区域维修",
                "LLM 兜底",
                "模型语义匹配",
                "是",
                "是",
                "composite_project",
                "同时命中多个工程域：公共设施、道路工程",
                "道路工程/路面维修 | 停车交通/车位改造",
            ]
        )
        worksheet.append(
            [
                "历史复合工程样本",
                "防水工程",
                "外墙防水",
                "规则优先",
                "历史结果",
                "是",
                "否",
                "composite_project",
                "同时命中多个工程域：防水工程、外立面修缮",
                "外立面修缮/外墙粉刷翻新",
            ]
        )
        worksheet.append(
            [
                "历史同域多系统样本",
                "消防",
                "消火栓维修",
                "规则优先",
                "历史结果",
                "否",
                "否",
                "multi_system_same_domain",
                "",
                "",
            ]
        )

        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        response = self.client.post(
            "/api/analyze-excel",
            files={
                "file": (
                    "classified.xlsx",
                    output.getvalue(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(data["summary"]["total_records"], 4)
        self.assertEqual(data["summary"]["llm_method_count"], 1)
        self.assertEqual(data["summary"]["fallback_method_count"], 0)
        self.assertEqual(data["summary"]["composite_count"], 2)
        self.assertEqual(data["structure_counts"]["composite_project"], 2)
        self.assertEqual(data["structure_counts"]["multi_system_same_domain"], 1)
        self.assertEqual(data["summary"]["review_count"], 3)
        self.assertEqual(len(data["focus_samples"]), 3)
        self.assertEqual(
            data["focus_samples"][0]["composite_reason"],
            "同时命中多个工程域：公共设施、道路工程",
        )
        self.assertEqual(
            data["focus_samples"][0]["secondary_candidates"],
            ["道路工程/路面维修", "停车交通/车位改造"],
        )
        self.assertTrue(all(sample["needs_review"] for sample in data["focus_samples"]))
        self.assertEqual(data["focus_samples"][0]["method"], "LLM 辅助分类")


if __name__ == "__main__":
    unittest.main()
