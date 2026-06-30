#!/usr/bin/env python3
import argparse
import csv
import hashlib
import re
from pathlib import Path
from typing import Any

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
BATCH_SAMPLE_NAME = "cost_item_samples.xlsx"
DEDUP_REPORT_HEADERS = [
    "stable_sample_id",
    "kept_batch_id",
    "duplicate_batch_id",
    "file_name",
    "consultation_project_name",
    "sub_project_id",
    "cost_item_name",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="合并所有批次清单样本并按 stable_sample_id 去重")
    parser.add_argument("--input-dir", default="samples", help="批次样本根目录，默认 samples")
    parser.add_argument("--output", default="samples/cost_item_samples_all.xlsx", help="总样本输出 xlsx 路径")
    parser.add_argument("--overwrite", action="store_true", help="若总样本或去重报告已存在则覆盖")
    return parser.parse_args()


def resolve_path(raw_path: str) -> Path:
    return Path(raw_path).expanduser().resolve()


def dedup_report_path(output_path: Path) -> Path:
    return output_path.with_name(f"{output_path.stem}_dedup_report.csv")


def find_sample_files(input_dir: Path) -> list[Path]:
    if not input_dir.exists():
        raise ValueError(f"样本批次目录不存在: {input_dir}")
    if not input_dir.is_dir():
        raise ValueError(f"样本批次路径不是目录: {input_dir}")
    return sorted(input_dir.glob(f"*/{BATCH_SAMPLE_NAME}"), key=lambda path: str(path))


def validate_output_paths(output_path: Path, report_path: Path, overwrite: bool) -> None:
    for path in (output_path, report_path):
        if path.exists() and path.is_dir():
            raise ValueError(f"输出路径是目录，不是文件: {path}")
    existing = [path for path in (output_path, report_path) if path.exists()]
    if existing and not overwrite:
        formatted = ", ".join(str(path) for path in existing)
        raise ValueError(f"输出已存在，请加 --overwrite 或更换 --output: {formatted}")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    report_path.parent.mkdir(parents=True, exist_ok=True)


def norm(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip().replace("\u3000", " ")
    return re.sub(r"\s+", " ", text).strip()


def row_text(row: dict[str, Any], *headers: str) -> str:
    for header in headers:
        value = row.get(header)
        text = norm(value)
        if text:
            return text
    return ""


def stable_sample_id(row: dict[str, Any]) -> str:
    stable_text = "|".join(
        [
            row_text(row, "file_name"),
            row_text(row, "consultation_project_name"),
            row_text(row, "consultation_time"),
            row_text(row, "location"),
            row_text(row, "renovation_content"),
            row_text(row, "sub_project_id"),
            row_text(row, "seq"),
            row_text(row, "project_name", "cost_item_name"),
            row_text(row, "project_description"),
            row_text(row, "unit", "unit_normalized"),
            row_text(row, "quantity"),
        ]
    )
    return hashlib.sha1(stable_text.encode("utf-8")).hexdigest()


def load_sample_sheet(path: Path):
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if "samples" not in workbook.sheetnames:
        workbook.close()
        raise ValueError(f"样本文件缺少 samples sheet: {path}")
    return workbook, workbook["samples"]


def header_values(sheet) -> list[str]:
    try:
        raw_headers = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    except StopIteration as exc:
        raise ValueError("samples sheet 为空") from exc
    return [norm(value) for value in raw_headers]


def make_report_row(
    stable_id: str,
    kept_batch_id: str,
    duplicate_batch_id: str,
    duplicate_row: dict[str, Any],
) -> dict[str, str]:
    return {
        "stable_sample_id": stable_id,
        "kept_batch_id": kept_batch_id,
        "duplicate_batch_id": duplicate_batch_id,
        "file_name": row_text(duplicate_row, "file_name"),
        "consultation_project_name": row_text(duplicate_row, "consultation_project_name"),
        "sub_project_id": row_text(duplicate_row, "sub_project_id"),
        "cost_item_name": row_text(duplicate_row, "project_name", "cost_item_name"),
    }


def normalize_values(values: tuple[Any, ...], length: int) -> list[Any]:
    row_values = list(values[:length])
    if len(row_values) < length:
        row_values.extend([None] * (length - len(row_values)))
    return row_values


def merge_batches(input_dir: Path, output_path: Path, report_path: Path) -> tuple[int, int, int]:
    sample_files = find_sample_files(input_dir)
    if not sample_files:
        raise ValueError(f"未找到批次样本文件: {input_dir}/*/{BATCH_SAMPLE_NAME}")

    base_headers: list[str] | None = None
    for sample_file in sample_files:
        workbook, sheet = load_sample_sheet(sample_file)
        try:
            headers = header_values(sheet)
            if base_headers is None:
                base_headers = headers
            elif headers != base_headers:
                raise ValueError(f"样本表头不一致: {sample_file}")
        finally:
            workbook.close()

    assert base_headers is not None

    output_workbook = openpyxl.Workbook(write_only=True)
    output_sheet = output_workbook.create_sheet("samples")
    output_sheet.append([*base_headers, "batch_id", "stable_sample_id"])

    seen: dict[str, str] = {}
    report_rows: list[dict[str, str]] = []
    input_rows = 0
    output_rows = 0

    for sample_file in sample_files:
        batch_id = sample_file.parent.name
        workbook, sheet = load_sample_sheet(sample_file)
        try:
            for values in sheet.iter_rows(min_row=2, values_only=True):
                input_rows += 1
                row_values = normalize_values(values, len(base_headers))
                row = dict(zip(base_headers, row_values))
                stable_id = stable_sample_id(row)
                kept_batch_id = seen.get(stable_id)
                if kept_batch_id is not None:
                    report_rows.append(make_report_row(stable_id, kept_batch_id, batch_id, row))
                    continue
                seen[stable_id] = batch_id
                output_sheet.append([*row_values, batch_id, stable_id])
                output_rows += 1
        finally:
            workbook.close()

    output_workbook.save(output_path)
    with report_path.open("w", encoding="utf-8-sig", newline="") as report_file:
        writer = csv.DictWriter(report_file, fieldnames=DEDUP_REPORT_HEADERS)
        writer.writeheader()
        writer.writerows(report_rows)

    return input_rows, output_rows, len(report_rows)


def main() -> int:
    args = parse_args()
    input_dir = resolve_path(args.input_dir)
    output_path = resolve_path(args.output)
    report_path = dedup_report_path(output_path)

    try:
        validate_output_paths(output_path, report_path, args.overwrite)
        input_rows, output_rows, duplicate_rows = merge_batches(input_dir, output_path, report_path)
    except ValueError as exc:
        print(f"[ERROR] {exc}")
        return 1

    print(f"[DONE] input rows: {input_rows}")
    print(f"[DONE] output rows: {output_rows}")
    print(f"[DONE] duplicate rows: {duplicate_rows}")
    print(f"[DONE] output: {output_path}")
    print(f"[DONE] dedup report: {report_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
