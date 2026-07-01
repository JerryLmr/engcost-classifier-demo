#!/usr/bin/env python3
import argparse
import json
import re
import sys
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
BACKEND_DIR = ROOT / "backend"
if str(BACKEND_DIR) not in sys.path:
    sys.path.insert(0, str(BACKEND_DIR))

from classifier.unit_normalizer import normalize_unit  # noqa: E402


PROJECT_HEADERS = [
    "file_name",
    "工程名称",
    "project_name_text",
    "consultation_project_name",
    "renovation_content",
    "sub_project_id",
    "catalog_id",
    "一级分类",
    "二级分类",
    "维修状态",
    "标准对象",
    "是否复合工程",
    "复合目录",
    "是否紧急维修",
    "是否白蚁相关",
    "是否建议复核",
    "分类依据",
    "consultation_time",
    "location",
]

SAMPLE_HEADERS = [
    "file_name",
    "工程名称",
    "cost_item_name",
    "project_description",
    "一级分类",
    "二级分类",
    "维修状态",
    "unit_normalized",
    "quantity",
    "unit_price",
    "total_price",
    "labor_cost",
    "machinery_cost",
    "labor_unit_price",
    "machinery_unit_price",
    "project_name_text",
    "consultation_project_name",
    "renovation_content",
    "sub_project_id",
    "catalog_id",
    "标准对象",
    "是否复合工程",
    "复合目录",
    "是否紧急维修",
    "是否白蚁相关",
    "是否建议复核",
    "分类依据",
    "seq",
    "unit",
    "item_similarity_text",
    "item_context_text",
    "consultation_time",
    "location",
    "source_row_id",
    "item_row_id",
    "source_json",
]

PARSE_ERROR_HEADERS = [
    "source_row_id",
    "error_type",
    "error_message",
    "raw_sub_item_project_rows",
    "raw_row_summary",
]

NUMERIC_FIELDS = [
    "quantity",
    "unit_price",
    "total_price",
    "labor_cost",
    "machinery_cost",
]

EXCEL_SUFFIXES = {".xlsx", ".xlsm"}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="构建已审定清单级样本 Excel")
    parser.add_argument("input_path", help="工程分类后的 Excel 文件")
    parser.add_argument(
        "-o",
        "--output",
        required=True,
        help="输出清单样本 Excel 路径",
    )
    parser.add_argument(
        "--overwrite",
        action="store_true",
        help="若输出文件已存在则覆盖",
    )
    return parser.parse_args()


def cell_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


DATE_TEXT_RE = re.compile(r"^(\d{4})[-/](\d{1,2})[-/](\d{1,2})(?:\s+0{1,2}:0{2}:0{2}(?:\.0+)?)?$")


def consultation_time_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()

    text = str(value).strip()
    if not text:
        return ""

    match = DATE_TEXT_RE.fullmatch(text)
    if not match:
        return text

    year, month, day = (int(part) for part in match.groups())
    try:
        return date(year, month, day).isoformat()
    except ValueError:
        return text


def load_header_map_from_values(header_row: tuple[Any, ...]) -> dict[str, int]:
    header_map: dict[str, int] = {}
    for index, value in enumerate(header_row):
        header = cell_text(value)
        if header and header not in header_map:
            header_map[header] = index
    return header_map


def get_cell_from_row(
    row_values: tuple[Any, ...],
    header_map: dict[str, int],
    header: str,
) -> Any:
    index = header_map.get(header)
    if index is None:
        return None
    if index >= len(row_values):
        return None
    return row_values[index]


def get_project_header_text_from_row(
    row_values: tuple[Any, ...],
    header_map: dict[str, int],
    header: str,
) -> str:
    value = get_cell_from_row(row_values, header_map, header)
    if header == "consultation_time":
        return consultation_time_text(value)
    return cell_text(value)


def build_project_name(project_name: str, consultation_project_name: str, renovation_content: str) -> str:
    if project_name:
        return project_name
    return " ".join(part for part in [consultation_project_name, renovation_content] if part).strip()


def compact_json(value: Any) -> str:
    try:
        return json.dumps(value, ensure_ascii=False, sort_keys=True)
    except TypeError:
        return str(value)


def raw_row_summary(row_values: dict[str, str]) -> str:
    summary = {
        "file_name": row_values.get("file_name", ""),
        "工程名称": row_values.get("工程名称", ""),
        "project_name_text": row_values.get("project_name_text", ""),
        "consultation_project_name": row_values.get("consultation_project_name", ""),
        "renovation_content": row_values.get("renovation_content", ""),
        "catalog_id": row_values.get("catalog_id", ""),
        "一级分类": row_values.get("一级分类", ""),
        "二级分类": row_values.get("二级分类", ""),
    }
    return compact_json(summary)


def make_error(
    source_row_id: int,
    error_type: str,
    error_message: str,
    raw_sub_item_project_rows: Any,
    row_summary: str,
) -> dict[str, Any]:
    return {
        "source_row_id": source_row_id,
        "error_type": error_type,
        "error_message": error_message,
        "raw_sub_item_project_rows": cell_text(raw_sub_item_project_rows),
        "raw_row_summary": row_summary,
    }


def parse_sub_item_rows(raw_value: Any) -> tuple[list[Any] | None, str | None]:
    if raw_value is None or cell_text(raw_value) == "":
        return None, "sub_item_project_rows 为空"
    if isinstance(raw_value, list):
        return raw_value, None
    if isinstance(raw_value, tuple):
        return list(raw_value), None
    if not isinstance(raw_value, str):
        return None, f"sub_item_project_rows 不是 JSON 字符串或数组: {type(raw_value).__name__}"

    raw_text = raw_value.strip()
    try:
        parsed = json.loads(raw_text)
    except json.JSONDecodeError as exc:
        return None, f"JSON 解析失败: {exc.msg} at pos {exc.pos}"

    if not isinstance(parsed, list):
        return None, f"sub_item_project_rows 解析后不是数组: {type(parsed).__name__}"
    return parsed, None


def parse_number(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip()
    if not text:
        return None

    normalized = (
        text.replace(",", "")
        .replace("，", "")
        .replace("￥", "")
        .replace("¥", "")
        .replace("$", "")
        .replace("元", "")
        .strip()
    )
    normalized = re.sub(r"\s+", "", normalized)
    if not normalized:
        return None

    try:
        return float(Decimal(normalized))
    except (InvalidOperation, ValueError):
        return None


def divide_or_none(numerator: float | None, denominator: float | None) -> float | None:
    if numerator is None or denominator is None or denominator == 0:
        return None
    return numerator / denominator


def clean_text(value: Any) -> str:
    text = cell_text(value)
    if not text:
        return ""

    lines: list[str] = []
    for line in re.split(r"[\r\n]+", text):
        stripped = line.strip()
        stripped = re.sub(r"^\s*\d+\s*(?:、|\.(?!\d))\s*", "", stripped)
        if stripped:
            lines.append(stripped)

    cleaned = " ".join(lines)
    cleaned = re.sub(r"(?:(?<=\s)|^)\d+\s*(?:、|\.(?!\d))\s*", "", cleaned)
    cleaned = re.sub(r"(?i)(\d+(?:\.\d+)?\s*mm)(sbs)", r"\1 \2", cleaned)
    cleaned = re.sub(r"\s+", " ", cleaned).strip()
    return cleaned


def collect_numeric_parse_errors(
    source_row_id: int,
    seq_text: str,
    item: dict[str, Any],
    numeric_values: dict[str, float | None],
    raw_sub_item_rows: Any,
    row_summary: str,
) -> list[dict[str, Any]]:
    errors: list[dict[str, Any]] = []
    for field, parsed_value in numeric_values.items():
        raw_value = item.get(field)
        if parsed_value is None and raw_value is not None and cell_text(raw_value) != "":
            errors.append(
                make_error(
                    source_row_id,
                    "invalid_numeric_field",
                    f"数值字段无法解析: seq={seq_text}, field={field}, value={raw_value}",
                    raw_sub_item_rows,
                    row_summary,
                )
            )

    if numeric_values.get("unit_price") is None:
        errors.append(
            make_error(
                source_row_id,
                "missing_unit_price",
                f"清单行缺少或无法解析 unit_price: seq={seq_text}",
                raw_sub_item_rows,
                row_summary,
            )
        )
    return errors


def join_text(parts: list[str], separator: str) -> str:
    return separator.join(part for part in parts if part)


def append_dict_row(worksheet, headers: list[str], row: dict[str, Any]) -> None:
    worksheet.append([row.get(header) for header in headers])


def append_error(
    errors_sheet,
    source_row_id: int,
    error_type: str,
    error_message: str,
    raw_sub_item_rows: Any,
    row_summary: str,
) -> None:
    append_dict_row(
        errors_sheet,
        PARSE_ERROR_HEADERS,
        make_error(
            source_row_id,
            error_type,
            error_message,
            raw_sub_item_rows,
            row_summary,
        ),
    )


def build_and_write_samples(input_path: Path, output_path: Path) -> tuple[int, int]:
    input_workbook = openpyxl.load_workbook(input_path, read_only=True, data_only=True)
    try:
        worksheet = input_workbook.active
        header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
        header_map = load_header_map_from_values(header_row)

        missing_required = [
            header for header in ("sub_item_project_rows", "project_name_text") if header not in header_map
        ]
        if missing_required:
            raise ValueError(f"输入 Excel 缺少必要列: {', '.join(missing_required)}")

        output_workbook = openpyxl.Workbook(write_only=True)
        samples_sheet = output_workbook.create_sheet("samples")
        errors_sheet = output_workbook.create_sheet("parse_errors")
        samples_sheet.append(SAMPLE_HEADERS)
        errors_sheet.append(PARSE_ERROR_HEADERS)

        samples_count = 0
        errors_count = 0

        for source_row_id, row_values_tuple in enumerate(
            worksheet.iter_rows(min_row=2, values_only=True),
            start=2,
        ):
            raw_sub_item_rows = get_cell_from_row(row_values_tuple, header_map, "sub_item_project_rows")
            row_values = {
                header: get_project_header_text_from_row(row_values_tuple, header_map, header)
                for header in PROJECT_HEADERS
            }
            row_values["工程名称"] = build_project_name(
                row_values["工程名称"],
                row_values["consultation_project_name"],
                row_values["renovation_content"],
            )
            row_summary = raw_row_summary(row_values)

            item_rows, parse_error = parse_sub_item_rows(raw_sub_item_rows)
            if parse_error:
                append_error(
                    errors_sheet,
                    source_row_id,
                    "invalid_sub_item_project_rows",
                    parse_error,
                    raw_sub_item_rows,
                    row_summary,
                )
                errors_count += 1
                continue

            if not item_rows:
                append_error(
                    errors_sheet,
                    source_row_id,
                    "empty_sub_item_project_rows",
                    "sub_item_project_rows 数组为空",
                    raw_sub_item_rows,
                    row_summary,
                )
                errors_count += 1
                continue

            for index, item in enumerate(item_rows, start=1):
                if not isinstance(item, dict):
                    append_error(
                        errors_sheet,
                        source_row_id,
                        "invalid_item_row",
                        f"清单行不是对象: index={index}, type={type(item).__name__}",
                        raw_sub_item_rows,
                        row_summary,
                    )
                    errors_count += 1
                    continue

                seq = item.get("seq")
                if seq is None or cell_text(seq) == "":
                    seq = index
                    append_error(
                        errors_sheet,
                        source_row_id,
                        "missing_seq",
                        f"清单行缺少 seq，已使用数组内序号: {index}",
                        raw_sub_item_rows,
                        row_summary,
                    )
                    errors_count += 1
                seq_text = cell_text(seq)

                cost_item_name = cell_text(item.get("cost_item_name") or item.get("project_name"))
                if not cost_item_name:
                    append_error(
                        errors_sheet,
                        source_row_id,
                        "missing_cost_item_name",
                        f"清单行缺少 cost_item_name/project_name: seq={seq_text}",
                        raw_sub_item_rows,
                        row_summary,
                    )
                    errors_count += 1

                project_description = cell_text(item.get("project_description"))
                cleaned_description = clean_text(project_description)
                unit = cell_text(item.get("unit"))
                numeric_values = {field: parse_number(item.get(field)) for field in NUMERIC_FIELDS}
                for error in collect_numeric_parse_errors(
                    source_row_id,
                    seq_text,
                    item,
                    numeric_values,
                    raw_sub_item_rows,
                    row_summary,
                ):
                    append_dict_row(errors_sheet, PARSE_ERROR_HEADERS, error)
                    errors_count += 1

                sample = {
                    "source_row_id": source_row_id,
                    "item_row_id": f"{source_row_id}-{seq_text}",
                    **row_values,
                    "sub_project_id": cell_text(item.get("sub_project_id")) or row_values.get("sub_project_id", ""),
                    "seq": seq,
                    "cost_item_name": cost_item_name,
                    "project_description": project_description,
                    "unit": unit,
                    "unit_normalized": normalize_unit(unit),
                    "quantity": numeric_values["quantity"],
                    "unit_price": numeric_values["unit_price"],
                    "total_price": numeric_values["total_price"],
                    "labor_cost": numeric_values["labor_cost"],
                    "machinery_cost": numeric_values["machinery_cost"],
                    "labor_unit_price": divide_or_none(numeric_values["labor_cost"], numeric_values["quantity"]),
                    "machinery_unit_price": divide_or_none(
                        numeric_values["machinery_cost"],
                        numeric_values["quantity"],
                    ),
                    "item_similarity_text": join_text([cost_item_name, cleaned_description], "；"),
                    "item_context_text": join_text(
                        [
                            row_values["工程名称"],
                            cell_text(item.get("sub_project_id")),
                            cost_item_name,
                            cleaned_description,
                        ],
                        " / ",
                    ),
                    "source_json": compact_json(item),
                }
                append_dict_row(samples_sheet, SAMPLE_HEADERS, sample)
                samples_count += 1

        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_workbook.save(output_path)
        return samples_count, errors_count
    finally:
        input_workbook.close()


def validate_paths(input_path: Path, output_path: Path, overwrite: bool) -> None:
    if not input_path.exists():
        raise ValueError(f"输入文件不存在: {input_path}")
    if not input_path.is_file():
        raise ValueError(f"输入路径不是文件: {input_path}")
    if input_path.suffix.lower() not in EXCEL_SUFFIXES:
        raise ValueError(f"输入文件不是 Excel: {input_path}")
    if output_path.exists() and output_path.is_dir():
        raise ValueError(f"输出路径是目录，不是文件: {output_path}")
    if output_path.exists() and not overwrite:
        raise ValueError(f"输出已存在，请加 --overwrite 或更换输出路径: {output_path}")


def main() -> int:
    args = parse_args()
    input_path = Path(args.input_path).expanduser().resolve()
    output_path = Path(args.output).expanduser().resolve()

    try:
        validate_paths(input_path, output_path, args.overwrite)
        samples_count, errors_count = build_and_write_samples(input_path, output_path)
    except ValueError as exc:
        print(f"[ERROR] {exc}")
        return 1

    print(f"[DONE] 输入文件: {input_path}")
    print(f"[DONE] 输出文件: {output_path}")
    print(f"[DONE] samples: {samples_count}")
    print(f"[DONE] parse_errors: {errors_count}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
