#!/usr/bin/env python3
import argparse
import csv
from pathlib import Path

from openpyxl import load_workbook


COMPARE_COLUMNS = ["一级分类", "二级分类", "分类方式", "是否复合工程", "是否建议复核", "结构类型"]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="对比两套 Excel 分类结果目录")
    parser.add_argument("left_dir", help="左侧结果目录，例如 excel_outputs_python")
    parser.add_argument("right_dir", help="右侧结果目录，例如 excel_outputs_json")
    parser.add_argument("--csv", help="可选：输出差异明细 CSV 路径")
    return parser.parse_args()


def read_rows(path: Path):
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    rows = worksheet.iter_rows(values_only=True)
    header = list(next(rows))
    index = {name: i for i, name in enumerate(header)}
    data = []
    for row_num, row in enumerate(rows, start=2):
        if not any(row):
            continue
        record = {"row_num": row_num, "工程名称": row[0] if row else ""}
        for column in COMPARE_COLUMNS:
            record[column] = row[index[column]] if column in index else None
        data.append(record)
    return data


def main() -> int:
    args = parse_args()
    left_dir = Path(args.left_dir).expanduser().resolve()
    right_dir = Path(args.right_dir).expanduser().resolve()

    left_files = {path.name: path for path in left_dir.glob("*.xlsx") if not path.name.startswith("~$")}
    right_files = {path.name: path for path in right_dir.glob("*.xlsx") if not path.name.startswith("~$")}

    missing_left = sorted(set(right_files) - set(left_files))
    missing_right = sorted(set(left_files) - set(right_files))
    common_files = sorted(set(left_files) & set(right_files))

    print(f"[INFO] 左侧目录: {left_dir}")
    print(f"[INFO] 右侧目录: {right_dir}")
    print(f"[INFO] 共同文件数: {len(common_files)}")

    diff_rows = []

    for filename in common_files:
        left_rows = read_rows(left_files[filename])
        right_rows = read_rows(right_files[filename])
        max_len = max(len(left_rows), len(right_rows))
        for idx in range(max_len):
            left = left_rows[idx] if idx < len(left_rows) else None
            right = right_rows[idx] if idx < len(right_rows) else None
            if left is None or right is None:
                diff_rows.append(
                    {
                        "file": filename,
                        "row_num": left["row_num"] if left else right["row_num"],
                        "工程名称": left["工程名称"] if left else right["工程名称"],
                        "column": "__row_presence__",
                        "left": left,
                        "right": right,
                    }
                )
                continue
            for column in COMPARE_COLUMNS:
                if left[column] != right[column]:
                    diff_rows.append(
                        {
                            "file": filename,
                            "row_num": left["row_num"],
                            "工程名称": left["工程名称"],
                            "column": column,
                            "left": left[column],
                            "right": right[column],
                        }
                    )

        print(f"[FILE] {filename}: 差异 {sum(1 for row in diff_rows if row['file'] == filename)} 条")

    print(f"[DONE] 左侧缺失文件: {len(missing_left)}")
    for name in missing_left:
        print(f"  - only_right: {name}")
    print(f"[DONE] 右侧缺失文件: {len(missing_right)}")
    for name in missing_right:
        print(f"  - only_left: {name}")

    print(f"[DONE] 字段差异总数: {len(diff_rows)}")

    if args.csv:
        csv_path = Path(args.csv).expanduser().resolve()
        csv_path.parent.mkdir(parents=True, exist_ok=True)
        with csv_path.open("w", encoding="utf-8-sig", newline="") as fp:
            writer = csv.DictWriter(fp, fieldnames=["file", "row_num", "工程名称", "column", "left", "right"])
            writer.writeheader()
            writer.writerows(diff_rows)
        print(f"[DONE] 差异明细已写入: {csv_path}")

    return 1 if diff_rows or missing_left or missing_right else 0


if __name__ == "__main__":
    raise SystemExit(main())
