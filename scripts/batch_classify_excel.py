#!/usr/bin/env python3
import argparse
import hashlib
import io
import json
import re
import sys
from datetime import date, datetime
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
BACKEND_DIR = ROOT / "backend"
if str(BACKEND_DIR) not in sys.path:
    sys.path.insert(0, str(BACKEND_DIR))


RESULT_HEADERS = [
    "工程名称",
    "project_name_text",
    "catalog_id",
    "一级分类",
    "二级分类",
    "维修状态",
    "标准对象",
    "是否复合工程",
    "复合目录",
    "是否紧急维修",
    "是否白蚁相关",
    "是否建议复核",
    "分类依据",
    "file_name",
    "consultation_project_name",
    "renovation_content",
    "sub_project_id",
    "sub_item_project_rows",
    "consultation_time",
    "location",
    "cache_subject",
]

CLASSIFICATION_CACHE_VERSION = "classification_by_sub_project_v1"

EXCLUDE_EXACT_ITEM_NAMES = {
    "脚手架搭拆",
    "安全文明施工费",
    "夜间施工增加费",
    "二次搬运费",
    "冬雨季施工增加费",
    "已完工程及设备保护费",
    "规费",
    "税金",
}

EXCLUDE_NAME_KEYWORDS = [
    "措施",
    "规费",
    "税金",
    "合计",
    "汇总",
    "暂列金额",
    "暂估价",
]

EXCEL_SUFFIXES = {".xlsx", ".xlsm"}
OCR_HEADERS = [
    "file_name",
    "consultation_project_name",
    "consultation_time",
    "renovation_content",
    "sub_item_project_rows",
    "location",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="批量处理 Excel 工程标准目录分类文件")
    parser.add_argument("input_paths", nargs="+", help="待处理 Excel 文件；单个目录也可批量处理")
    parser.add_argument(
        "-o",
        "--output",
        help="输出文件或目录。多个输入或目录输入时按输出目录处理",
    )
    parser.add_argument(
        "--overwrite",
        action="store_true",
        help="若输出文件已存在则覆盖",
    )
    parser.add_argument(
        "--include-classified",
        action="store_true",
        help="目录模式默认跳过已带 _分类结果 / _classified 后缀的文件；设置后不跳过",
    )
    return parser.parse_args()


def should_skip_file(path: Path, include_classified: bool) -> bool:
    if path.suffix.lower() not in EXCEL_SUFFIXES:
        return True
    if include_classified:
        return False
    stem = path.stem
    return stem.endswith("_分类结果") or stem.endswith("_classified")


def classified_output_name(input_file: Path) -> str:
    return f"{input_file.stem}_classified.xlsx"


def _bool_text(value: object) -> str:
    return "是" if bool(value) else "否"


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def norm_text(value: object) -> str:
    s = "" if value is None else str(value)
    return re.sub(r"\s+", "", s).strip()


def normalize_for_cache_subject(value: object) -> str:
    s = norm_text(value)
    if not s:
        return ""

    s = re.sub(r"\d+", "", s)
    s = re.sub(r"[#＃_\-－—~～/／\\.,，、:：;；()（）\[\]【】{}<>《》+*×$]", "", s)
    return s.strip()


def _classification_cache_subject(unit_project_name: object) -> str:
    return normalize_for_cache_subject(unit_project_name)


def _display_match_text(value: object) -> str:
    return re.sub(r"[-_－—]+", "", norm_text(value))


def _build_unit_project_name(consultation_project_name: object, sub_project_id: object) -> str:
    consultation_name = _cell_text(consultation_project_name)
    subject = _cell_text(sub_project_id)

    if consultation_name and subject:
        if subject.startswith(consultation_name):
            return subject
        consultation_base = _display_match_text(consultation_name)
        subject_base = _display_match_text(subject)
        if consultation_base and subject_base and (
            subject_base.startswith(consultation_base) or consultation_base.startswith(subject_base)
        ):
            return subject
        return f"{consultation_name}-{subject}"

    return consultation_name or subject


def clean_sub_project_id(value: object, project_code: object | None = None) -> str:
    s = "" if value is None else str(value)
    s = re.sub(r"\s+", "", s.strip())

    if project_code:
        code = re.escape(str(project_code).strip())
        s = re.sub(rf"[-_－—]?{code}$", "", s)

    s = re.sub(r"[-_－—]\d{9,12}$", "", s)
    return s.strip("-_－— ")


def _strip_trailing_project_codes(value: object, project_codes: set[str]) -> str:
    s = norm_text(value)
    if not s:
        return ""

    codes = sorted(
        {norm_text(code) for code in project_codes if norm_text(code)},
        key=len,
        reverse=True,
    )
    if not codes:
        return s

    changed = True
    while changed:
        changed = False
        for code in codes:
            for sep in ("-", "_", "－", "—"):
                suffix = sep + code
                if s.endswith(suffix):
                    s = s[: -len(suffix)]
                    s = s.strip("-_－— ")
                    changed = True
                    break
            if changed:
                break

    return s.strip("-_－— ")


def is_classifiable_item(item: dict) -> bool:
    name = str(item.get("project_name") or "").strip()
    code = str(item.get("project_code") or "").strip()

    if not name:
        return False
    if name in EXCLUDE_EXACT_ITEM_NAMES:
        return False
    if any(k in name for k in EXCLUDE_NAME_KEYWORDS):
        return False
    if name == "脚手架搭拆" and str(item.get("unit_price") or "").strip() == code:
        return False
    return True


_DATE_TEXT_RE = re.compile(r"^(\d{4})[-/](\d{1,2})[-/](\d{1,2})(?:\s+0{1,2}:0{2}:0{2}(?:\.0+)?)?$")


def _consultation_time_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()

    text = str(value).strip()
    if not text:
        return ""

    match = _DATE_TEXT_RE.fullmatch(text)
    if not match:
        return text

    year, month, day = (int(part) for part in match.groups())
    try:
        return date(year, month, day).isoformat()
    except ValueError:
        return text


def _load_header_map(worksheet) -> dict[str, int]:
    header_map: dict[str, int] = {}
    for column in range(1, worksheet.max_column + 1):
        header = _cell_text(worksheet.cell(row=1, column=column).value)
        if header and header not in header_map:
            header_map[header] = column
    return header_map


def _get_cell(worksheet, header_map: dict[str, int], row: int, header: str) -> object:
    column = header_map.get(header)
    if not column:
        return None
    return worksheet.cell(row=row, column=column).value


def _get_text(worksheet, header_map: dict[str, int], row: int, header: str) -> str:
    return _cell_text(_get_cell(worksheet, header_map, row, header))


def _build_project_name(consultation_project_name: object, renovation_content: object) -> str:
    return " ".join(
        part for part in [_cell_text(consultation_project_name), _cell_text(renovation_content)] if part
    ).strip()


def _parse_sub_item_project_rows(raw_value: object) -> list[dict[str, object]]:
    if raw_value is None or _cell_text(raw_value) == "":
        return []
    if isinstance(raw_value, list):
        parsed = raw_value
    elif isinstance(raw_value, tuple):
        parsed = list(raw_value)
    elif isinstance(raw_value, str):
        try:
            parsed = json.loads(raw_value.strip())
        except json.JSONDecodeError:
            return []
    else:
        return []

    if not isinstance(parsed, list):
        return []
    return [dict(item) for item in parsed if isinstance(item, dict)]


def _json_dumps(value: object) -> str:
    return json.dumps(value, ensure_ascii=False, sort_keys=True)


def _merge_key(ocr_values: dict[str, object]) -> tuple[str, str, str, str, str]:
    return (
        norm_text(ocr_values.get("file_name")),
        norm_text(ocr_values.get("consultation_project_name")),
        norm_text(ocr_values.get("consultation_time")),
        norm_text(ocr_values.get("location")),
        norm_text(ocr_values.get("renovation_content")),
    )


def _dedupe_key(item: dict[str, object]) -> tuple[str, str, str, str, str]:
    return (
        str(item.get("page_no", "")),
        norm_text(item.get("seq")),
        norm_text(item.get("project_code")),
        norm_text(item.get("project_name")),
        norm_text(item.get("total_price")),
    )


def _sort_value(value: object) -> tuple[int, float | str]:
    text = norm_text(value)
    if not text:
        return (1, "")
    try:
        return (0, float(text))
    except ValueError:
        return (1, text)


def _sort_key(item: dict[str, object]) -> tuple[tuple[int, float | str], tuple[int, float | str], str, str]:
    return (
        _sort_value(item.get("page_no")),
        _sort_value(item.get("seq")),
        norm_text(item.get("project_code")),
        norm_text(item.get("project_name")),
    )


def _prepare_merged_items(items: list[dict[str, object]], fallback_subject: str) -> list[dict[str, object]]:
    prepared: list[dict[str, object]] = []
    seen: set[tuple[str, str, str, str, str]] = set()
    project_codes = {
        norm_text(item.get("project_code"))
        for item in items
        if norm_text(item.get("project_code"))
    }
    for item in sorted(items, key=_sort_key):
        cleaned = clean_sub_project_id(item.get("sub_project_id"), item.get("project_code"))
        cleaned = _strip_trailing_project_codes(cleaned, project_codes)
        updated = dict(item)
        updated["sub_project_id"] = cleaned or fallback_subject or "未分组"
        key = _dedupe_key(updated)
        if key in seen:
            continue
        seen.add(key)
        prepared.append(updated)
    return prepared


def _item_name(item: dict[str, object]) -> str:
    return _cell_text(item.get("project_name") or item.get("cost_item_name"))


def _group_classification_units(
    ocr_values: dict[str, object],
    items: list[dict[str, object]],
) -> list[dict[str, object]]:
    fallback_subject = _cell_text(ocr_values.get("consultation_project_name")) or "未分组"
    prepared_items = _prepare_merged_items(items, fallback_subject)
    if not prepared_items:
        unit_project_name = _build_unit_project_name(
            ocr_values.get("consultation_project_name"),
            fallback_subject,
        )
        return [
            {
                "ocr_values": {
                    **ocr_values,
                    "sub_project_id": fallback_subject,
                    "sub_item_project_rows": "[]",
                    "unit_project_name": unit_project_name,
                },
                "classify_subject": fallback_subject,
                "unit_project_name": unit_project_name,
                "item_names": [],
            }
        ]

    grouped_items: dict[str, list[dict[str, object]]] = {}
    filtered_names: dict[str, list[str]] = {}
    for item in prepared_items:
        subject = _cell_text(item.get("sub_project_id")) or fallback_subject
        grouped_items.setdefault(subject, []).append(item)
        if is_classifiable_item(item):
            name = _item_name(item)
            if name:
                filtered_names.setdefault(subject, []).append(name)

    units: list[dict[str, object]] = []
    for subject, group_items in grouped_items.items():
        unit_ocr_values = dict(ocr_values)
        unit_ocr_values["sub_project_id"] = subject
        unit_ocr_values["sub_item_project_rows"] = _json_dumps(group_items)
        unit_project_name = _build_unit_project_name(
            unit_ocr_values.get("consultation_project_name"),
            subject,
        )
        unit_ocr_values["unit_project_name"] = unit_project_name
        units.append(
            {
                "ocr_values": unit_ocr_values,
                "classify_subject": subject,
                "unit_project_name": unit_project_name,
                "item_names": filtered_names.get(subject, []),
            }
        )
    return units


def _validate_input_headers(header_map: dict[str, int], first_header: object) -> None:
    missing_ocr_headers = [header for header in OCR_HEADERS if header not in header_map]
    if not missing_ocr_headers:
        return

    if first_header is None or str(first_header).strip() == "":
        raise ValueError("第一列表头不能为空")
    raise ValueError(
        "新版 OCR 输入 Excel 必须包含字段: "
        + ", ".join(OCR_HEADERS)
        + f"。当前缺少: {', '.join(missing_ocr_headers)}"
    )


def _read_ocr_values(worksheet, header_map: dict[str, int], row: int) -> dict[str, object]:
    values = {header: _get_cell(worksheet, header_map, row, header) for header in OCR_HEADERS}
    values["consultation_time"] = _consultation_time_text(values.get("consultation_time"))
    return values


def _write_result_row(
    worksheet,
    row: int,
    result: dict[str, object],
    ocr_values: dict[str, object],
) -> None:
    values = [
        (
            ocr_values.get("unit_project_name")
            or ocr_values.get("sub_project_id")
            or ocr_values.get("consultation_project_name")
        ),
        result.get("project_name_text", ""),
        result.get("catalog_id", ""),
        result.get("category", ""),
        result.get("item", ""),
        result.get("repair_status", ""),
        result.get("standard_group", ""),
        _bool_text(result.get("is_composite")) if "is_composite" in result else "",
        " | ".join(result.get("secondary_catalog_labels") or []),
        _bool_text(result.get("is_emergency")) if "is_emergency" in result else "",
        _bool_text(result.get("termite_related")) if "termite_related" in result else "",
        _bool_text(result.get("needs_review")) if "needs_review" in result else "",
        result.get("reason", ""),
        ocr_values.get("file_name"),
        ocr_values.get("consultation_project_name"),
        ocr_values.get("renovation_content"),
        ocr_values.get("sub_project_id"),
        ocr_values.get("sub_item_project_rows"),
        ocr_values.get("consultation_time"),
        ocr_values.get("location"),
        ocr_values.get("cache_subject"),
    ]
    for column, value in enumerate(values, start=1):
        worksheet.cell(row=row, column=column, value=value)


def _classification_cache_key(
    consultation_project_name: object,
    renovation_content: object,
    unit_project_name: object,
) -> str:
    cache_subject = _classification_cache_subject(unit_project_name)
    raw_key = (
        norm_text(consultation_project_name)
        + "|"
        + norm_text(renovation_content)
        + "|"
        + cache_subject
    )
    return f"{CLASSIFICATION_CACHE_VERSION}:{hashlib.sha256(raw_key.encode('utf-8')).hexdigest()}"


def _ensure_project_name_text(result: dict[str, object], project_text: str, path: Path, source_row: int) -> dict[str, object]:
    project_name_text = _cell_text(result.get("project_name_text"))
    if project_name_text:
        return result

    print(
        f"[WARN] {path.name}:{source_row} classification result missing project_name_text; "
        "fallback to original 工程名称",
        flush=True,
    )
    updated = dict(result)
    updated["project_name_text"] = project_text
    return updated


def classify_workbook(
    path: Path,
    output_path: Path,
    classify_project_func,
    classification_cache: dict[str, dict[str, object]],
) -> tuple[int, int, int, int]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    source_sheet = workbook.active

    first_header = source_sheet.cell(row=1, column=1).value
    header_map = _load_header_map(source_sheet)
    _validate_input_headers(header_map, first_header)

    output_workbook = openpyxl.Workbook()
    output_sheet = output_workbook.active
    output_sheet.title = "分类结果"
    for column, header in enumerate(RESULT_HEADERS, start=1):
        output_sheet.cell(row=1, column=column, value=header)

    processed = 0
    classify_call_count = 0
    cache_hit_count = 0
    empty_project_name_rows = 0
    output_rows = 0
    output_row = 2
    merged_projects: dict[tuple[str, str, str, str, str], dict[str, object]] = {}
    for source_row in range(2, source_sheet.max_row + 1):
        ocr_values = _read_ocr_values(source_sheet, header_map, source_row)
        merge_key = _merge_key(ocr_values)
        if merge_key not in merged_projects:
            merged_projects[merge_key] = {
                "ocr_values": ocr_values,
                "items": [],
                "source_rows": [],
            }
        merged_projects[merge_key]["items"].extend(_parse_sub_item_project_rows(ocr_values.get("sub_item_project_rows")))
        merged_projects[merge_key]["source_rows"].append(source_row)

    for merged_project in merged_projects.values():
        base_ocr_values = merged_project["ocr_values"]
        items = merged_project["items"]
        source_rows = merged_project["source_rows"]
        first_source_row = source_rows[0] if source_rows else 0
        units = _group_classification_units(base_ocr_values, items)

        for unit in units:
            ocr_values = unit["ocr_values"]
            classify_subject = _cell_text(unit["classify_subject"])
            unit_project_name = _cell_text(ocr_values.get("unit_project_name") or unit.get("unit_project_name"))
            item_names = list(unit["item_names"])
            if classify_subject == "未分组":
                empty_project_name_rows += 1

            cache_subject = _classification_cache_subject(unit_project_name)
            ocr_values["cache_subject"] = cache_subject
            cache_key = _classification_cache_key(
                ocr_values.get("consultation_project_name"),
                ocr_values.get("renovation_content"),
                unit_project_name,
            )
            if cache_key in classification_cache:
                result = classification_cache[cache_key]
                cache_hit_count += 1
                print(
                    f"[CACHE] {path.name}:{first_source_row} "
                    f"{unit_project_name[:80]} cache_subject={cache_subject[:80]}",
                    flush=True,
                )
            else:
                print(
                    f"[ROW ] {path.name}:{first_source_row} "
                    f"{unit_project_name[:80]} cache_subject={cache_subject[:80]}",
                    flush=True,
                )

                result = classify_project_func(
                    classify_subject,
                    consultation_project_name=_cell_text(ocr_values.get("consultation_project_name")),
                    item_summary=item_names,
                )
                if result.get("pipeline_status") == "llm_service_error":
                    raise RuntimeError(f"LLM 服务连接失败，已停止处理当前文件。失败行: {first_source_row}")
                result = _ensure_project_name_text(result, classify_subject, path, first_source_row)
                classification_cache[cache_key] = result
                classify_call_count += 1

                print(
                    f"[DONE] {path.name}:{first_source_row} "
                    f"{result.get('catalog_id')} "
                    f"{result.get('category')} / {result.get('item')} "
                    f"status={result.get('pipeline_status')}",
                    flush=True,
                )

            _write_result_row(output_sheet, output_row, result, ocr_values)
            output_row += 1
            output_rows += 1
            processed += 1

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output = io.BytesIO()
    output_workbook.save(output)
    output_path.write_bytes(output.getvalue())
    workbook.close()
    return classify_call_count, cache_hit_count, output_rows, empty_project_name_rows


def _is_directory_output(raw_output: str, output_path: Path) -> bool:
    return (
        raw_output.endswith(("/", "\\"))
        or output_path.exists() and output_path.is_dir()
        or (not output_path.exists() and output_path.suffix == "")
    )


def resolve_single_file_output(input_path: Path, output_arg: str | None) -> Path:
    if not output_arg:
        return input_path.with_name(classified_output_name(input_path)).resolve()

    raw_output = output_arg.strip()
    output_path = Path(raw_output).expanduser()
    if _is_directory_output(raw_output, output_path):
        return (output_path / classified_output_name(input_path)).resolve()
    return output_path.resolve()


def _directory_jobs(input_dir: Path, output_arg: str | None, include_classified: bool) -> list[tuple[Path, Path]]:
    output_dir = Path(output_arg).expanduser().resolve() if output_arg else input_dir / "classified_results"
    excel_files = sorted(path for path in input_dir.iterdir() if not should_skip_file(path, include_classified))
    return [(path, output_dir / classified_output_name(path)) for path in excel_files]


def resolve_jobs(input_paths: list[str], output_arg: str | None, include_classified: bool) -> list[tuple[Path, Path]]:
    paths = [Path(raw).expanduser().resolve() for raw in input_paths]
    if len(paths) == 1 and paths[0].is_dir():
        return _directory_jobs(paths[0], output_arg, include_classified)

    if any(path.is_dir() for path in paths):
        raise ValueError("多个输入路径时不支持混用目录；请传多个 Excel 文件或单个目录")

    if len(paths) == 1:
        return [(paths[0], resolve_single_file_output(paths[0], output_arg))]

    if output_arg:
        output_dir = Path(output_arg).expanduser().resolve()
        if output_dir.suffix:
            raise ValueError("多个输入文件时 -o/--output 必须是目录")
    else:
        output_dir = paths[0].parent
    return [(path, output_dir / classified_output_name(path)) for path in paths]


def validate_jobs(jobs: list[tuple[Path, Path]], overwrite: bool) -> None:
    for input_path, output_path in jobs:
        if not input_path.exists():
            raise ValueError(f"输入文件不存在: {input_path}")
        if not input_path.is_file():
            raise ValueError(f"输入路径不是文件: {input_path}")
        if input_path.suffix.lower() not in EXCEL_SUFFIXES:
            raise ValueError(f"输入文件不是 Excel: {input_path}")
        if output_path.exists() and output_path.is_dir():
            raise ValueError(f"输出路径是目录，不是文件: {output_path}")
        if output_path.exists() and not overwrite:
            raise ValueError(f"输出已存在，请加 --overwrite 或更换输出路径: {output_path}")
        output_path.parent.mkdir(parents=True, exist_ok=True)


def main() -> int:
    from classifier.llm_client import check_lmstudio_service  # noqa: E402
    from services.standard_classifier import classify_project_standard  # noqa: E402

    args = parse_args()
    try:
        jobs = resolve_jobs(args.input_paths, args.output, args.include_classified)
        validate_jobs(jobs, args.overwrite)
    except ValueError as exc:
        print(f"[ERROR] {exc}")
        return 1

    if not jobs:
        print("[ERROR] 没有可处理的 Excel 文件")
        return 1

    try:
        check_lmstudio_service()
    except RuntimeError as exc:
        print(f"[ERROR] {exc}")
        return 1

    total_files = 0
    total_processed = 0
    total_classify_calls = 0
    total_cache_hits = 0
    total_output_rows = 0
    total_skipped = 0
    failed_files: list[tuple[str, str]] = []
    classification_cache: dict[str, dict[str, object]] = {}

    print(f"[INFO] 待处理文件数: {len(jobs)}")
    for path, output_path in jobs:
        print(f"[PLAN] {path} -> {output_path}")

    for path, output_path in jobs:
        print(f"[RUN ] {path.name}")
        try:
            (
                classify_call_count,
                cache_hit_count,
                output_rows,
                empty_project_name_rows,
            ) = classify_workbook(path, output_path, classify_project_standard, classification_cache)
            total_files += 1
            processed = classify_call_count + cache_hit_count
            total_processed += processed
            total_classify_calls += classify_call_count
            total_cache_hits += cache_hit_count
            total_output_rows += output_rows
            total_skipped += empty_project_name_rows
            print(
                f"[ OK ] {path.name} -> {output_path.name} "
                f"(分类 {processed} 行, 空工程名保留 {empty_project_name_rows} 行)"
            )
        except Exception as exc:  # noqa: BLE001
            failed_files.append((path.name, str(exc)))
            print(f"[FAIL] {path.name}: {exc}")

    print()
    print(f"[DONE] 成功文件: {total_files}, 失败文件: {len(failed_files)}")
    print(f"[DONE] 总分类行数: {total_processed}, 空工程名保留行数: {total_skipped}")
    print(f"[DONE] 实际分类调用次数: {total_classify_calls}")
    print(f"[DONE] 缓存命中次数: {total_cache_hits}")
    print(f"[DONE] 输出行数: {total_output_rows}")
    if failed_files:
        print("[DONE] 失败明细:")
        for name, error in failed_files:
            print(f"  - {name}: {error}")
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
