#!/usr/bin/env python3
import argparse
import math
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl


REQUIRED_HEADERS = [
    "file_name",
    "consultation_project_name",
    "consultation_time",
    "renovation_content",
    "sub_item_project_rows",
    "location",
]

REMOVED_PREFIX_HEADERS = [
    "source_row_id",
    "missing_required_fields",
    "removed_reason",
]

EXCEL_SUFFIXES = {".xlsx", ".xlsm"}
EMPTY_TEXT_VALUES = {"null", "none", "nan"}
REMOVED_REASON = "缺少必填 OCR 字段"
_DATE_TEXT_RE = re.compile(
    r"^(\d{4})[-/](\d{1,2})[-/](\d{1,2})(?:\s+\d{1,2}:\d{2}:\d{2}(?:\.0+)?)?$"
)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="过滤 OCR Excel 中缺少必填字段的行")
    parser.add_argument("input_excel", help="OCR 输入 Excel 文件")
    parser.add_argument("--clean-output", required=True, help="必填字段完整行输出 xlsx 路径")
    parser.add_argument("--removed-output", required=True, help="缺少必填字段行输出 xlsx 路径")
    parser.add_argument("--overwrite", action="store_true", help="若输出文件已存在则覆盖")
    return parser.parse_args()


def cell_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_consultation_time(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()

    text = str(value).strip()
    if not text:
        return ""

    text = text.replace("年", "-").replace("月", "-").replace("日", "")
    text = text.replace("/", "-")

    match = _DATE_TEXT_RE.fullmatch(text)
    if not match:
        return text

    year, month, day = (int(part) for part in match.groups())
    try:
        return date(year, month, day).isoformat()
    except ValueError:
        return text


def normalize_consultation_time_in_row(row_values: list[Any], header_map: dict[str, int]) -> list[Any]:
    normalized_row_values = list(row_values)
    time_column = header_map.get("consultation_time")
    if time_column is not None and time_column < len(normalized_row_values):
        normalized_row_values[time_column] = normalize_consultation_time(normalized_row_values[time_column])
    return normalized_row_values


def is_empty_required_value(value: object) -> bool:
    if value is None:
        return True
    if isinstance(value, float) and math.isnan(value):
        return True
    if isinstance(value, str):
        text = value.strip()
        return not text or text.lower() in EMPTY_TEXT_VALUES
    return False


def resolve_path(raw_path: str) -> Path:
    return Path(raw_path).expanduser().resolve()


def validate_paths(input_path: Path, clean_output: Path, removed_output: Path, overwrite: bool) -> None:
    if not input_path.exists():
        raise ValueError(f"输入文件不存在: {input_path}")
    if not input_path.is_file():
        raise ValueError(f"输入路径不是文件: {input_path}")
    if input_path.suffix.lower() not in EXCEL_SUFFIXES:
        raise ValueError(f"输入文件不是 Excel: {input_path}")

    if clean_output == input_path or removed_output == input_path:
        raise ValueError("输出文件不能覆盖原始输入文件")
    if clean_output == removed_output:
        raise ValueError("clean-output 和 removed-output 不能是同一个文件")

    for output_path in (clean_output, removed_output):
        if output_path.exists() and output_path.is_dir():
            raise ValueError(f"输出路径是目录，不是文件: {output_path}")
        if output_path.exists() and not overwrite:
            raise ValueError(f"输出已存在，请加 --overwrite 或更换输出路径: {output_path}")
        output_path.parent.mkdir(parents=True, exist_ok=True)


def build_header_map(headers: list[Any]) -> dict[str, int]:
    header_map: dict[str, int] = {}
    for index, raw_header in enumerate(headers):
        header = cell_text(raw_header)
        if header and header not in header_map:
            header_map[header] = index
    return header_map


def validate_required_headers(header_map: dict[str, int]) -> None:
    missing_headers = [header for header in REQUIRED_HEADERS if header not in header_map]
    if missing_headers:
        raise ValueError(
            "输入 Excel 缺少必要列: "
            + ", ".join(missing_headers)
            + "。必须包含: "
            + ", ".join(REQUIRED_HEADERS)
        )


def normalize_row_length(row_values: list[Any], expected_length: int) -> list[Any]:
    if len(row_values) >= expected_length:
        return row_values[:expected_length]
    return row_values + [None] * (expected_length - len(row_values))


def filter_required_ocr_rows(
    input_path: Path,
    clean_output: Path,
    removed_output: Path,
) -> tuple[int, int, int, dict[str, int]]:
    source_workbook = openpyxl.load_workbook(input_path, read_only=True, data_only=True)
    source_sheet = source_workbook.active

    row_iter = source_sheet.iter_rows(values_only=True)

    try:
        headers = list(next(row_iter))
    except StopIteration as exc:
        source_workbook.close()
        raise ValueError("输入 Excel 为空") from exc

    header_map = build_header_map(headers)
    validate_required_headers(header_map)

    expected_length = len(headers)
    missing_counts = {header: 0 for header in REQUIRED_HEADERS}

    clean_workbook = openpyxl.Workbook(write_only=True)
    clean_sheet = clean_workbook.create_sheet("cleaned")
    clean_sheet.append(headers)

    removed_workbook = openpyxl.Workbook(write_only=True)
    removed_sheet = removed_workbook.create_sheet("removed")
    removed_sheet.append(REMOVED_PREFIX_HEADERS + headers)

    input_rows = 0
    cleaned_rows = 0
    removed_rows = 0

    try:
        for source_row_id, row in enumerate(row_iter, start=2):
            row_values = normalize_row_length(list(row), expected_length)
            normalized_row_values = normalize_consultation_time_in_row(row_values, header_map)
            input_rows += 1

            missing_fields: list[str] = []
            for header in REQUIRED_HEADERS:
                column_index = header_map[header]
                value = row_values[column_index] if column_index < len(row_values) else None
                if is_empty_required_value(value):
                    missing_fields.append(header)

            if missing_fields:
                for field in missing_fields:
                    missing_counts[field] += 1

                removed_sheet.append(
                    [
                        source_row_id,
                        ",".join(missing_fields),
                        REMOVED_REASON,
                        *normalized_row_values,
                    ]
                )
                removed_rows += 1
            else:
                clean_sheet.append(normalized_row_values)
                cleaned_rows += 1

        clean_workbook.save(clean_output)
        removed_workbook.save(removed_output)
    finally:
        source_workbook.close()

    return input_rows, cleaned_rows, removed_rows, missing_counts


def print_summary(input_rows: int, cleaned_rows: int, removed_rows: int, missing_counts: dict[str, int]) -> None:
    print(f"[DONE] input rows: {input_rows}")
    print(f"[DONE] cleaned rows: {cleaned_rows}")
    print(f"[DONE] removed rows: {removed_rows}")
    print("[DONE] missing counts:")
    for header in REQUIRED_HEADERS:
        print(f"  {header}: {missing_counts[header]}")


def main() -> int:
    args = parse_args()
    input_path = resolve_path(args.input_excel)
    clean_output = resolve_path(args.clean_output)
    removed_output = resolve_path(args.removed_output)

    try:
        validate_paths(input_path, clean_output, removed_output, args.overwrite)
        input_rows, cleaned_rows, removed_rows, missing_counts = filter_required_ocr_rows(
            input_path,
            clean_output,
            removed_output,
        )
    except Exception as exc:  # noqa: BLE001
        print(f"[ERROR] {exc}")
        return 1

    print_summary(input_rows, cleaned_rows, removed_rows, missing_counts)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
